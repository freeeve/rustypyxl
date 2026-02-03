"""Tests for style/formatting preservation during load → save roundtrip.

These tests verify that formatting is preserved when opening and resaving files.
Bug report: "formatting getting removed after opening and resaving a file"
"""

import pytest
import rustypyxl


class TestFontRoundtrip:
    """Tests for font style preservation."""

    def test_font_bold_preserved(self, tmp_path):
        """Bold font should survive roundtrip."""
        # Create file with bold font
        path1 = tmp_path / "bold.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Bold Text"
        ws["A1"].font = rustypyxl.Font(bold=True)
        wb.save(str(path1))

        # Reload and resave
        path2 = tmp_path / "bold_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        # Load resaved and check
        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.bold is True

    def test_font_italic_preserved(self, tmp_path):
        """Italic font should survive roundtrip."""
        path1 = tmp_path / "italic.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Italic Text"
        ws["A1"].font = rustypyxl.Font(italic=True)
        wb.save(str(path1))

        path2 = tmp_path / "italic_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.italic is True

    def test_font_size_preserved(self, tmp_path):
        """Font size should survive roundtrip."""
        path1 = tmp_path / "size.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Large Text"
        ws["A1"].font = rustypyxl.Font(size=24)
        wb.save(str(path1))

        path2 = tmp_path / "size_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.size == 24

    def test_font_color_preserved(self, tmp_path):
        """Font color should survive roundtrip."""
        path1 = tmp_path / "color.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Red Text"
        ws["A1"].font = rustypyxl.Font(color="FF0000")
        wb.save(str(path1))

        path2 = tmp_path / "color_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.color is not None
        # Color might be stored as ARGB (FFFF0000) or RGB (FF0000)
        assert "FF0000" in wb3["Test"]["A1"].font.color.upper()

    def test_font_name_preserved(self, tmp_path):
        """Font name should survive roundtrip."""
        path1 = tmp_path / "fontname.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Arial Text"
        ws["A1"].font = rustypyxl.Font(name="Arial")
        wb.save(str(path1))

        path2 = tmp_path / "fontname_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.name == "Arial"

    def test_font_strike_preserved(self, tmp_path):
        """Strikethrough font should survive roundtrip."""
        path1 = tmp_path / "strike.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Strikethrough Text"
        ws["A1"].font = rustypyxl.Font(strike=True)
        wb.save(str(path1))

        path2 = tmp_path / "strike_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.strike is True

    def test_font_superscript_preserved(self, tmp_path):
        """Superscript vertAlign should survive roundtrip."""
        path1 = tmp_path / "superscript.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Superscript Text"
        ws["A1"].font = rustypyxl.Font(vertAlign="superscript")
        wb.save(str(path1))

        path2 = tmp_path / "superscript_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.vertAlign == "superscript"

    def test_font_subscript_preserved(self, tmp_path):
        """Subscript vertAlign should survive roundtrip."""
        path1 = tmp_path / "subscript.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Subscript Text"
        ws["A1"].font = rustypyxl.Font(vertAlign="subscript")
        wb.save(str(path1))

        path2 = tmp_path / "subscript_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.vertAlign == "subscript"


class TestFillRoundtrip:
    """Tests for fill/background color preservation."""

    def test_solid_fill_preserved(self, tmp_path):
        """Solid fill color should survive roundtrip."""
        path1 = tmp_path / "fill.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Yellow Background"
        ws["A1"].fill = rustypyxl.PatternFill(
            fill_type="solid",
            fgColor="FFFF00"
        )
        wb.save(str(path1))

        path2 = tmp_path / "fill_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].fill is not None
        assert wb3["Test"]["A1"].fill.fill_type == "solid"
        assert "FFFF00" in wb3["Test"]["A1"].fill.fgColor.upper()


class TestBorderRoundtrip:
    """Tests for border preservation."""

    def test_thin_border_preserved(self, tmp_path):
        """Thin border should survive roundtrip."""
        path1 = tmp_path / "border.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Bordered"
        ws["A1"].border = rustypyxl.Border(
            left=rustypyxl.Side(style="thin"),
            right=rustypyxl.Side(style="thin"),
            top=rustypyxl.Side(style="thin"),
            bottom=rustypyxl.Side(style="thin"),
        )
        wb.save(str(path1))

        path2 = tmp_path / "border_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        border = wb3["Test"]["A1"].border
        assert border is not None
        assert border.left.style == "thin"
        assert border.right.style == "thin"
        assert border.top.style == "thin"
        assert border.bottom.style == "thin"

    def test_colored_border_preserved(self, tmp_path):
        """Colored border should survive roundtrip."""
        path1 = tmp_path / "colored_border.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Red Border"
        ws["A1"].border = rustypyxl.Border(
            left=rustypyxl.Side(style="thick", color="FF0000"),
        )
        wb.save(str(path1))

        path2 = tmp_path / "colored_border_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        border = wb3["Test"]["A1"].border
        assert border is not None
        assert border.left.style == "thick"
        assert "FF0000" in border.left.color.upper()


class TestProtectionRoundtrip:
    """Tests for cell protection preservation."""

    def test_unlocked_cell_preserved(self, tmp_path):
        """Unlocked cell should survive roundtrip."""
        path1 = tmp_path / "unlocked.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Unlocked Cell"
        ws["A1"].protection = rustypyxl.Protection(locked=False)
        wb.save(str(path1))

        path2 = tmp_path / "unlocked_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        prot = wb3["Test"]["A1"].protection
        assert prot is not None
        assert prot.locked is False

    def test_hidden_cell_preserved(self, tmp_path):
        """Hidden formula cell should survive roundtrip."""
        path1 = tmp_path / "hidden.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Hidden Formula"
        ws["A1"].protection = rustypyxl.Protection(locked=True, hidden=True)
        wb.save(str(path1))

        path2 = tmp_path / "hidden_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        prot = wb3["Test"]["A1"].protection
        assert prot is not None
        assert prot.locked is True
        assert prot.hidden is True


class TestAlignmentRoundtrip:
    """Tests for alignment preservation."""

    def test_horizontal_alignment_preserved(self, tmp_path):
        """Horizontal alignment should survive roundtrip."""
        path1 = tmp_path / "halign.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Centered"
        ws["A1"].alignment = rustypyxl.Alignment(horizontal="center")
        wb.save(str(path1))

        path2 = tmp_path / "halign_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].alignment.horizontal == "center"

    def test_vertical_alignment_preserved(self, tmp_path):
        """Vertical alignment should survive roundtrip."""
        path1 = tmp_path / "valign.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Top"
        ws["A1"].alignment = rustypyxl.Alignment(vertical="top")
        wb.save(str(path1))

        path2 = tmp_path / "valign_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].alignment.vertical == "top"

    def test_wrap_text_preserved(self, tmp_path):
        """Wrap text should survive roundtrip."""
        path1 = tmp_path / "wrap.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Wrapped\nText"
        ws["A1"].alignment = rustypyxl.Alignment(wrap_text=True)
        wb.save(str(path1))

        path2 = tmp_path / "wrap_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].alignment.wrap_text is True

    def test_text_rotation_preserved(self, tmp_path):
        """Text rotation should survive roundtrip."""
        path1 = tmp_path / "rotation.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Rotated"
        ws["A1"].alignment = rustypyxl.Alignment(text_rotation=45)
        wb.save(str(path1))

        path2 = tmp_path / "rotation_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].alignment.text_rotation == 45


class TestNumberFormatRoundtrip:
    """Tests for number format preservation."""

    def test_currency_format_preserved(self, tmp_path):
        """Currency format should survive roundtrip."""
        path1 = tmp_path / "currency.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = 1234.56
        ws["A1"].number_format = "$#,##0.00"
        wb.save(str(path1))

        path2 = tmp_path / "currency_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].number_format == "$#,##0.00"

    def test_percentage_format_preserved(self, tmp_path):
        """Percentage format should survive roundtrip."""
        path1 = tmp_path / "percent.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = 0.75
        ws["A1"].number_format = "0.00%"
        wb.save(str(path1))

        path2 = tmp_path / "percent_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].number_format == "0.00%"

    def test_date_format_preserved(self, tmp_path):
        """Date format should survive roundtrip."""
        path1 = tmp_path / "date.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = 44927  # Excel date serial for 2023-01-01
        ws["A1"].number_format = "YYYY-MM-DD"
        wb.save(str(path1))

        path2 = tmp_path / "date_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].number_format == "YYYY-MM-DD"


class TestMultipleStyledCells:
    """Tests for multiple cells with different styles."""

    def test_different_fonts_same_sheet(self, tmp_path):
        """Multiple cells with different fonts should preserve all."""
        path1 = tmp_path / "multi_font.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")

        ws["A1"].value = "Bold"
        ws["A1"].font = rustypyxl.Font(bold=True)

        ws["A2"].value = "Italic"
        ws["A2"].font = rustypyxl.Font(italic=True)

        ws["A3"].value = "Large"
        ws["A3"].font = rustypyxl.Font(size=20)

        wb.save(str(path1))

        path2 = tmp_path / "multi_font_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].font.bold is True
        assert wb3["Test"]["A2"].font.italic is True
        assert wb3["Test"]["A3"].font.size == 20

    def test_combined_styles(self, tmp_path):
        """Cell with multiple style properties should preserve all."""
        path1 = tmp_path / "combined.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")

        ws["A1"].value = "Styled Cell"
        ws["A1"].font = rustypyxl.Font(bold=True, size=14, color="0000FF")
        ws["A1"].fill = rustypyxl.PatternFill(fill_type="solid", fgColor="FFFF00")
        ws["A1"].alignment = rustypyxl.Alignment(horizontal="center")

        wb.save(str(path1))

        path2 = tmp_path / "combined_resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        cell = wb3["Test"]["A1"]
        assert cell.font.bold is True
        assert cell.font.size == 14
        assert "0000FF" in cell.font.color.upper()
        assert cell.fill.fill_type == "solid"
        assert cell.alignment.horizontal == "center"


class TestExistingFileRoundtrip:
    """Tests using existing formatted xlsx files."""

    def test_load_save_preserves_formatting(self, tmp_path):
        """Loading and saving existing file should preserve formatting."""
        # First create a file with formatting using openpyxl-style API
        path1 = tmp_path / "original.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")

        # Header row with bold
        for col, header in enumerate(["Name", "Value", "Status"], 1):
            ws.cell(row=1, column=col).value = header
            ws.cell(row=1, column=col).font = rustypyxl.Font(bold=True)

        # Data rows
        ws.cell(row=2, column=1).value = "Item A"
        ws.cell(row=2, column=2).value = 100
        ws.cell(row=2, column=2).number_format = "#,##0"

        wb.save(str(path1))

        # Now load and save without modifications
        path2 = tmp_path / "resaved.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        # Verify formatting preserved
        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Data"].cell(row=1, column=1).font.bold is True
        assert wb3["Data"].cell(row=2, column=2).number_format == "#,##0"


class TestThemeColorRoundtrip:
    """Tests for theme color preservation in fills and borders."""

    def test_theme_fill_not_corrupted_to_black(self, tmp_path):
        """Theme-colored fills should not become black on roundtrip.

        Regression test: write_fill_xml was emitting <fgColor rgb="FFtheme:0"/>
        instead of <fgColor theme="0"/>, causing Excel to interpret as black.
        """
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
        except ImportError:
            pytest.skip("openpyxl not installed")

        # Create file with theme fill using openpyxl
        path1 = tmp_path / "theme_fill.xlsx"
        wb_openpyxl = openpyxl.Workbook()
        ws = wb_openpyxl.active
        ws["A1"].value = "Theme Fill"
        ws["A1"].fill = PatternFill(start_color="00FF00", fill_type="solid")
        ws["B1"].value = "Another Fill"
        ws["B1"].fill = PatternFill(start_color="0000FF", fill_type="solid")
        wb_openpyxl.save(str(path1))

        # Roundtrip through rustypyxl
        path2 = tmp_path / "theme_fill_roundtrip.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        # Reload with openpyxl and verify colors are not corrupted
        wb3 = openpyxl.load_workbook(str(path2))
        ws3 = wb3.active
        fill_a1 = ws3["A1"].fill
        assert fill_a1.fgColor is not None
        # The color should not contain "theme:" as a raw string in rgb
        fg_rgb = fill_a1.fgColor.rgb if fill_a1.fgColor.rgb else ""
        assert "theme" not in fg_rgb.lower(), (
            f"Theme color was corrupted to rgb value: {fg_rgb}"
        )

    def test_theme_fill_roundtrip_rustypyxl_only(self, tmp_path):
        """Fill with theme color string should roundtrip correctly in rustypyxl."""
        path1 = tmp_path / "fill_rgb.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Red Fill"
        ws["A1"].fill = rustypyxl.PatternFill(fill_type="solid", fgColor="FF0000")
        ws["B1"].value = "Green Fill"
        ws["B1"].fill = rustypyxl.PatternFill(fill_type="solid", fgColor="00FF00")
        wb.save(str(path1))

        # Roundtrip
        path2 = tmp_path / "fill_rgb_roundtrip.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        # Verify fills preserved
        wb3 = rustypyxl.load_workbook(str(path2))
        assert wb3["Test"]["A1"].fill is not None
        assert wb3["Test"]["A1"].fill.fill_type == "solid"
        fg_a1 = wb3["Test"]["A1"].fill.fgColor
        assert fg_a1 is not None
        assert "FF0000" in fg_a1.upper()
        assert "theme" not in fg_a1.lower()

        fg_b1 = wb3["Test"]["B1"].fill.fgColor
        assert fg_b1 is not None
        assert "00FF00" in fg_b1.upper()

    def test_colored_border_not_corrupted(self, tmp_path):
        """Borders with colors should not become black on roundtrip."""
        path1 = tmp_path / "border_color.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Test")
        ws["A1"].value = "Bordered"
        ws["A1"].border = rustypyxl.Border(
            left=rustypyxl.Side(style="thin", color="FF0000"),
            top=rustypyxl.Side(style="thin", color="00FF00"),
        )
        wb.save(str(path1))

        # Roundtrip
        path2 = tmp_path / "border_color_roundtrip.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        wb3 = rustypyxl.load_workbook(str(path2))
        border = wb3["Test"]["A1"].border
        assert border is not None
        assert border.left.style == "thin"
        assert "FF0000" in border.left.color.upper()
        assert border.top.style == "thin"
        assert "00FF00" in border.top.color.upper()


class TestMultiSheetRoundtrip:
    """Tests for multi-sheet workbook roundtrip preservation."""

    def test_five_sheets_preserved(self, tmp_path):
        """Workbook with 5+ sheets should preserve all sheets on roundtrip."""
        sheet_names = ["Alpha", "Beta", "Gamma", "Delta", "Epsilon", "Zeta"]
        path1 = tmp_path / "multi_sheet.xlsx"
        wb = rustypyxl.Workbook()
        for i, name in enumerate(sheet_names):
            ws = wb.create_sheet(name)
            for row in range(1, 4):
                ws.cell(row=row, column=1).value = f"{name}_R{row}"
                ws.cell(row=row, column=2).value = (i + 1) * row
        wb.save(str(path1))

        # Roundtrip
        path2 = tmp_path / "multi_sheet_roundtrip.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        # Verify all sheets present with correct data
        wb3 = rustypyxl.load_workbook(str(path2))
        assert len(wb3.sheetnames) == len(sheet_names)
        for i, name in enumerate(sheet_names):
            assert name in wb3.sheetnames, f"Sheet '{name}' missing after roundtrip"
            ws = wb3[name]
            for row in range(1, 4):
                assert ws.cell(row=row, column=1).value == f"{name}_R{row}"
                assert ws.cell(row=row, column=2).value == (i + 1) * row

    def test_multi_sheet_zip_structure(self, tmp_path):
        """Roundtripped multi-sheet workbook should have correct ZIP structure."""
        import zipfile

        sheet_names = ["Sheet1", "Sheet2", "Sheet3", "Sheet4", "Sheet5"]
        path1 = tmp_path / "zip_check.xlsx"
        wb = rustypyxl.Workbook()
        for name in sheet_names:
            ws = wb.create_sheet(name)
            ws["A1"].value = name
        wb.save(str(path1))

        # Roundtrip
        path2 = tmp_path / "zip_check_roundtrip.xlsx"
        wb2 = rustypyxl.load_workbook(str(path1))
        wb2.save(str(path2))

        # Inspect ZIP contents
        with zipfile.ZipFile(str(path2), "r") as zf:
            names = zf.namelist()
            # Verify all sheet XML files exist
            for i in range(1, len(sheet_names) + 1):
                sheet_path = f"xl/worksheets/sheet{i}.xml"
                assert sheet_path in names, f"{sheet_path} missing from ZIP"

            # Verify Content_Types lists all sheets
            content_types = zf.read("[Content_Types].xml").decode()
            for i in range(1, len(sheet_names) + 1):
                assert f"/xl/worksheets/sheet{i}.xml" in content_types

            # Verify workbook.xml.rels has all relationships
            rels = zf.read("xl/_rels/workbook.xml.rels").decode()
            for i in range(1, len(sheet_names) + 1):
                assert f"worksheets/sheet{i}.xml" in rels
