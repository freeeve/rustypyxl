"""Data-validation rules added through the Python binding, read back by
openpyxl.
"""

import openpyxl
import rustypyxl


def test_list_validation_round_trips(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws.add_data_validation("A1:A10", "list", formula1='"Red,Green,Blue"')

    dvs = ws.data_validations
    assert len(dvs) == 1
    assert dvs[0]["type"] == "list"
    assert dvs[0]["sqref"] == "A1:A10"

    out = str(tmp_path / "dv.xlsx")
    wb.save(out)

    ows = openpyxl.load_workbook(out)["S"]
    dv_list = list(ows.data_validations.dataValidation)
    assert len(dv_list) == 1
    assert dv_list[0].type == "list"
    assert "A1:A10" in str(dv_list[0].sqref)


def test_whole_number_between_validation(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws.add_data_validation(
        "B1:B5", "whole", operator="between", formula1="1", formula2="100"
    )
    out = str(tmp_path / "dv2.xlsx")
    wb.save(out)
    ows = openpyxl.load_workbook(out)["S"]
    dv = list(ows.data_validations.dataValidation)[0]
    assert dv.type == "whole"
    assert dv.operator == "between"
