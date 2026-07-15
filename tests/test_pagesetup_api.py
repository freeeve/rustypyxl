"""Page setup, margins, print area, and headers/footers set through the Python
binding, read back by openpyxl.
"""

import openpyxl
import rustypyxl


def test_page_setup_and_margins(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws["A1"] = "x"
    ws.set_page_setup(orientation="landscape", paper_size="A4", fit_to_width=1)
    ws.set_page_margins(left=1.0, right=1.0, top=0.5, bottom=0.5)
    out = str(tmp_path / "ps.xlsx")
    wb.save(out)

    ows = openpyxl.load_workbook(out)["S"]
    assert ows.page_setup.orientation == "landscape"
    assert abs(ows.page_margins.left - 1.0) < 1e-6


def test_print_area_property(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws["A1"] = "x"
    ws.print_area = "A1:D20"
    assert ws.print_area == "A1:D20"
    out = str(tmp_path / "pa.xlsx")
    wb.save(out)
    ows = openpyxl.load_workbook(out)["S"]
    # Excel stores the print area as an absolute, sheet-qualified defined name.
    assert "$A$1:$D$20" in str(ows.print_area)


def test_header_footer(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws["A1"] = "x"
    ws.set_header_footer(header_center="Report", footer_right="Page &P")
    out = str(tmp_path / "hf.xlsx")
    wb.save(out)
    ows = openpyxl.load_workbook(out)["S"]
    assert ows.oddHeader.center.text == "Report"
