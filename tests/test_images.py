"""Images embed through rustypyxl and round-trip: an image added via the API
opens in openpyxl, and an image in an openpyxl/Excel-authored file survives a
rustypyxl load->save.
"""

import io

import openpyxl
import pytest
import rustypyxl

# A minimal valid 1x1 PNG.
PNG_1X1 = bytes(
    [
        0x89, 0x50, 0x4E, 0x47, 0x0D, 0x0A, 0x1A, 0x0A, 0x00, 0x00, 0x00, 0x0D,
        0x49, 0x48, 0x44, 0x52, 0x00, 0x00, 0x00, 0x01, 0x00, 0x00, 0x00, 0x01,
        0x08, 0x06, 0x00, 0x00, 0x00, 0x1F, 0x15, 0xC4, 0x89, 0x00, 0x00, 0x00,
        0x0A, 0x49, 0x44, 0x41, 0x54, 0x78, 0x9C, 0x63, 0x00, 0x01, 0x00, 0x00,
        0x05, 0x00, 0x01, 0x0D, 0x0A, 0x2D, 0xB4, 0x00, 0x00, 0x00, 0x00, 0x49,
        0x45, 0x4E, 0x44, 0xAE, 0x42, 0x60, 0x82,
    ]
)


def test_add_image_from_bytes_opens_in_openpyxl(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws.add_image(PNG_1X1, anchor="B2", name="Logo")
    out = str(tmp_path / "img.xlsx")
    wb.save(out)

    loaded = openpyxl.load_workbook(out)["S"]
    assert len(loaded._images) == 1


def test_add_image_from_path(tmp_path):
    png = tmp_path / "pic.png"
    png.write_bytes(PNG_1X1)

    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws.add_image(str(png), anchor="A1", width=64, height=64)
    out = str(tmp_path / "img.xlsx")
    wb.save(out)

    assert len(openpyxl.load_workbook(out)["S"]._images) == 1


def test_openpyxl_image_survives_rustypyxl_round_trip(tmp_path):
    # Author a file with an image using openpyxl.
    from openpyxl.drawing.image import Image as OpenpyxlImage

    src = str(tmp_path / "src.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pics"
    ws.add_image(OpenpyxlImage(io.BytesIO(PNG_1X1)), "C3")
    wb.save(src)

    # Round-trip it through rustypyxl.
    out = str(tmp_path / "out.xlsx")
    rustypyxl.load_workbook(src).save(out)

    # openpyxl still finds the image.
    assert len(openpyxl.load_workbook(out)["Pics"]._images) == 1


def test_bad_image_bytes_raise(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    with pytest.raises(ValueError):
        ws.add_image(b"not an image", anchor="A1")
