"""Column and row dimensions match openpyxl's idiom:
ws.column_dimensions['A'].width and ws.row_dimensions[1].height, and they
survive a save/load round-trip.
"""

import openpyxl
import rustypyxl


def test_column_and_row_dimensions_round_trip(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws["A1"] = "x"
    ws.column_dimensions["A"].width = 25.0
    ws.column_dimensions["C"].width = 12.5
    ws.row_dimensions[1].height = 30.0

    # Read back through the proxy.
    assert ws.column_dimensions["A"].width == 25.0
    assert ws.column_dimensions["C"].width == 12.5
    assert ws.row_dimensions[1].height == 30.0
    assert ws.column_dimensions["B"].width is None  # unset

    out = str(tmp_path / "dims.xlsx")
    wb.save(out)

    # openpyxl reads the same widths/heights.
    ows = openpyxl.load_workbook(out)["S"]
    assert ows.column_dimensions["A"].width == 25.0
    assert abs(ows.row_dimensions[1].height - 30.0) < 1e-6


def test_dimension_index_accessors():
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    assert ws.column_dimensions["D"].index == "D"
    assert ws.row_dimensions[7].index == 7
