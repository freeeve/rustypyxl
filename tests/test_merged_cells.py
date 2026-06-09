"""Tests for merged cell support.

Merges are verified by saving and reopening with openpyxl, so a no-op
implementation cannot pass.
"""

import io

import openpyxl
import rustypyxl


def merged_ranges_via_openpyxl(wb, sheet_name):
    data = wb.save_to_bytes()
    op = openpyxl.load_workbook(io.BytesIO(data))[sheet_name]
    return {str(r) for r in op.merged_cells.ranges}


class TestMergedCellBasics:
    """Test basic merged cell operations."""

    def test_merged_cells_empty_default(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        assert ws.merged_cells == []

    def test_merge_cells_by_range_string(self, workbook_with_sheet):
        wb = workbook_with_sheet
        ws = wb.active
        ws["A1"] = "merged"
        ws.merge_cells("A1:B2")
        assert ws.merged_cells == ["A1:B2"]
        assert merged_ranges_via_openpyxl(wb, "Test") == {"A1:B2"}

    def test_merge_cells_by_coordinates(self, workbook_with_sheet):
        wb = workbook_with_sheet
        ws = wb.active
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=3)
        assert ws.merged_cells == ["A1:C2"]
        assert merged_ranges_via_openpyxl(wb, "Test") == {"A1:C2"}

    def test_unmerge_cells(self, workbook_with_sheet):
        wb = workbook_with_sheet
        ws = wb.active
        ws.merge_cells("A1:B2")
        ws.unmerge_cells("A1:B2")
        assert ws.merged_cells == []
        assert merged_ranges_via_openpyxl(wb, "Test") == set()

    def test_multiple_merges(self, workbook_with_sheet):
        wb = workbook_with_sheet
        ws = wb.active
        ws.merge_cells("A1:B1")
        ws.merge_cells("A3:C3")
        assert set(ws.merged_cells) == {"A1:B1", "A3:C3"}
        assert merged_ranges_via_openpyxl(wb, "Test") == {"A1:B1", "A3:C3"}
