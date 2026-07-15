"""Excel tables added through the Python ws.add_table binding, read back by
openpyxl (the core supported tables; this exercises the new pyo3 binding).
"""

import openpyxl
import rustypyxl


def test_add_table_round_trips(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    wb.write_rows("S", [["Name", "Qty"], ["Widget", 3], ["Gadget", 5]])
    ws.add_table("Inventory", "A1:B3", style="TableStyleMedium9")

    assert [t["name"] for t in ws.tables] == ["Inventory"]
    assert ws.tables[0]["ref"] == "A1:B3"

    out = str(tmp_path / "tables.xlsx")
    wb.save(out)

    ows = openpyxl.load_workbook(out)["S"]
    assert "Inventory" in ows.tables
    assert ows.tables["Inventory"].ref == "A1:B3"


def test_add_table_with_options(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    wb.write_rows("S", [["A", "B"], [1, 2]])
    ws.add_table("T", "A1:B2", totals_row=False, row_stripes=True, auto_filter=True)
    out = str(tmp_path / "t.xlsx")
    wb.save(out)
    assert "T" in openpyxl.load_workbook(out)["S"].tables
