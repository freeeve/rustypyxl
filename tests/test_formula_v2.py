"""Expanded formula coverage and calculate_all write-back.

calculate_all evaluates every formula cell and stores its result as the cached
value, so a saved file shows computed results (openpyxl reads those cached
values back).
"""

import openpyxl
import rustypyxl


def _wb():
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    wb.write_rows("S", [[1], [2], [3], [4]])  # A1..A4
    return wb


def test_expanded_functions():
    wb = _wb()
    assert wb.evaluate_formula("S", "=MEDIAN(A1:A4)") == 2.5
    assert wb.evaluate_formula("S", "=ROUNDUP(2.1,0)") == 3
    assert wb.evaluate_formula("S", "=CEILING(7,5)") == 10
    assert wb.evaluate_formula("S", "=IFERROR(1/0,99)") == 99
    assert wb.evaluate_formula("S", '=SUBSTITUTE("a-b","-","+")') == "a+b"
    assert wb.evaluate_formula("S", '=PROPER("hello world")') == "Hello World"
    assert wb.evaluate_formula("S", "=DATE(2023,1,15)") == 44941
    assert wb.evaluate_formula("S", "=YEAR(44941)") == 2023


def test_lookup_over_real_cells():
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    wb.write_rows("S", [[1, "one"], [2, "two"], [3, "three"]])
    assert wb.evaluate_formula("S", "=VLOOKUP(2,A1:B3,2,FALSE)") == "two"
    assert wb.evaluate_formula("S", "=INDEX(A1:B3,3,2)") == "three"
    assert wb.evaluate_formula("S", "=MATCH(3,A1:A3,0)") == 3


def test_calculate_all_persists_cached_values(tmp_path):
    wb = _wb()
    ws = wb["S"]
    ws["B1"] = "=SUM(A1:A4)"
    ws["B2"] = "=B1*10"
    count = wb.calculate_all()
    assert count == 2

    out = str(tmp_path / "calc.xlsx")
    wb.save(out)

    # openpyxl (data_only) reads Excel's cached values -- which we computed.
    loaded = openpyxl.load_workbook(out, data_only=True)["S"]
    assert loaded["B1"].value == 10
    assert loaded["B2"].value == 100
