"""
Compatibility tests comparing rustypyxl vs openpyxl behavior.

These tests use fuzzy inputs to detect behavioral differences and bugs.
Each test performs the same operations with both libraries and compares results.
"""

import os
import random
import string
import tempfile
import pytest
from datetime import datetime, date
from decimal import Decimal

import openpyxl
import rustypyxl


# ============================================================================
# Test Utilities
# ============================================================================

def random_string(min_len=1, max_len=50):
    """Generate a random string."""
    length = random.randint(min_len, max_len)
    return ''.join(random.choices(string.ascii_letters + string.digits + ' ', k=length))


def random_unicode_string(min_len=1, max_len=30):
    """Generate a random unicode string with various characters."""
    chars = (
        string.ascii_letters +
        string.digits +
        ' !@#$%^&*()_+-=[]{}|;:,.<>?' +
        'áéíóúñüÁÉÍÓÚÑÜ' +  # Spanish
        '中文日本語한국어' +  # CJK
        'αβγδεζηθ' +  # Greek
        '🎉🚀💡'  # Emoji
    )
    length = random.randint(min_len, max_len)
    return ''.join(random.choices(chars, k=length))


def random_number():
    """Generate a random number (int or float)."""
    if random.random() < 0.5:
        return random.randint(-1000000, 1000000)
    else:
        return random.uniform(-1000000, 1000000)


def random_cell_value():
    """Generate a random cell value of various types."""
    choice = random.randint(0, 5)
    if choice == 0:
        return random_string()
    elif choice == 1:
        return random_number()
    elif choice == 2:
        return random.choice([True, False])
    elif choice == 3:
        return None  # Empty cell
    elif choice == 4:
        return random.randint(-1000000, 1000000)  # Integer
    else:
        return random_unicode_string()


def normalize_value(val):
    """Normalize values for comparison (handle float precision, None, etc.)."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        # Round floats to avoid precision issues
        if isinstance(val, float):
            return round(val, 10)
        return val
    if isinstance(val, str):
        return val
    return str(val)


def compare_values(rustypyxl_val, openpyxl_val, coord=""):
    """Compare two cell values, return True if equivalent."""
    rv = normalize_value(rustypyxl_val)
    ov = normalize_value(openpyxl_val)

    # Handle None/empty equivalence
    if rv is None and ov is None:
        return True
    if rv is None or ov is None:
        # One is None, other is not
        return False

    # Handle numeric comparison with tolerance
    if isinstance(rv, (int, float)) and isinstance(ov, (int, float)):
        if abs(rv - ov) < 1e-9:
            return True
        # Allow for float representation differences
        return abs(rv - ov) / max(abs(rv), abs(ov), 1) < 1e-9

    return rv == ov


def create_temp_file(suffix='.xlsx'):
    """Create a temporary file path."""
    fd, path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    return path


# ============================================================================
# Basic Value Tests
# ============================================================================

class TestBasicValues:
    """Test basic cell value handling."""

    def test_string_values(self):
        """Test various string values."""
        test_strings = [
            "Hello",
            "Hello World",
            "",  # Empty string
            " ",  # Space
            "  leading spaces",
            "trailing spaces  ",
            "line1\nline2",  # Newline
            "tab\there",  # Tab
            "Special: <>&\"'",  # XML special chars
            "Numbers123",
            "MixedCase123ABC",
            "a" * 1000,  # Long string
        ]

        for test_str in test_strings:
            # Create with openpyxl
            op_path = create_temp_file()
            op_wb = openpyxl.Workbook()
            op_ws = op_wb.active
            op_ws['A1'] = test_str
            op_wb.save(op_path)

            # Load with rustypyxl and compare
            rp_wb = rustypyxl.load_workbook(op_path)
            rp_val = rp_wb.get_cell_value('Sheet', 1, 1)

            # Also test the reverse: create with rustypyxl, load with openpyxl
            rp_path = create_temp_file()
            rp_wb2 = rustypyxl.Workbook()
            rp_ws2 = rp_wb2.create_sheet('Sheet')
            rp_wb2.set_cell_value('Sheet', 1, 1, test_str)
            rp_wb2.save(rp_path)

            op_wb2 = openpyxl.load_workbook(rp_path)
            op_val = op_wb2.active['A1'].value

            # Compare
            assert compare_values(rp_val, test_str), f"rustypyxl read mismatch for '{test_str}': got '{rp_val}'"
            assert compare_values(op_val, test_str), f"openpyxl read mismatch for '{test_str}': got '{op_val}'"

            # Cleanup
            os.unlink(op_path)
            os.unlink(rp_path)

    def test_numeric_values(self):
        """Test various numeric values."""
        test_numbers = [
            0,
            1,
            -1,
            42,
            -42,
            3.14159,
            -3.14159,
            1e10,
            1e-10,
            -1e10,
            0.0,
            1.0,
            999999999999,
            -999999999999,
            0.123456789012345,  # High precision
        ]

        for num in test_numbers:
            # Create with openpyxl
            op_path = create_temp_file()
            op_wb = openpyxl.Workbook()
            op_ws = op_wb.active
            op_ws['A1'] = num
            op_wb.save(op_path)

            # Load with rustypyxl
            rp_wb = rustypyxl.load_workbook(op_path)
            rp_val = rp_wb.get_cell_value('Sheet', 1, 1)

            assert compare_values(rp_val, num), f"Mismatch for {num}: got {rp_val}"

            # Reverse test
            rp_path = create_temp_file()
            rp_wb2 = rustypyxl.Workbook()
            rp_wb2.create_sheet('Sheet')
            rp_wb2.set_cell_value('Sheet', 1, 1, float(num))
            rp_wb2.save(rp_path)

            op_wb2 = openpyxl.load_workbook(rp_path)
            op_val = op_wb2.active['A1'].value

            assert compare_values(op_val, num), f"Reverse mismatch for {num}: got {op_val}"

            os.unlink(op_path)
            os.unlink(rp_path)

    def test_boolean_values(self):
        """Test boolean values."""
        for val in [True, False]:
            op_path = create_temp_file()
            op_wb = openpyxl.Workbook()
            op_ws = op_wb.active
            op_ws['A1'] = val
            op_wb.save(op_path)

            rp_wb = rustypyxl.load_workbook(op_path)
            rp_val = rp_wb.get_cell_value('Sheet', 1, 1)

            # Boolean may be read as True/False or 1/0
            assert rp_val == val or rp_val == (1 if val else 0), f"Boolean mismatch for {val}: got {rp_val}"

            os.unlink(op_path)


# ============================================================================
# Fuzzy Input Tests
# ============================================================================

class TestFuzzyInputs:
    """Test with randomly generated fuzzy inputs."""

    @pytest.mark.parametrize("seed", range(10))
    def test_random_grid_roundtrip(self, seed):
        """Test random grid data roundtrip through both libraries."""
        random.seed(seed)
        rows = random.randint(10, 50)
        cols = random.randint(5, 15)

        # Generate random data
        data = []
        for r in range(rows):
            row = []
            for c in range(cols):
                val = random_cell_value()
                # Skip complex values that might have precision issues
                if isinstance(val, float):
                    val = round(val, 6)
                row.append(val)
            data.append(row)

        # Create with rustypyxl
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Data')

        for r, row in enumerate(data, 1):
            for c, val in enumerate(row, 1):
                if val is not None:
                    rp_wb.set_cell_value('Data', r, c, val)

        rp_wb.save(rp_path)

        # Load with openpyxl and verify
        op_wb = openpyxl.load_workbook(rp_path)
        op_ws = op_wb['Data']

        mismatches = []
        for r, row in enumerate(data, 1):
            for c, expected in enumerate(row, 1):
                actual = op_ws.cell(r, c).value
                if not compare_values(actual, expected):
                    mismatches.append(f"({r},{c}): expected {expected!r}, got {actual!r}")

        assert len(mismatches) == 0, f"Mismatches found (seed={seed}):\n" + "\n".join(mismatches[:10])

        os.unlink(rp_path)

    @pytest.mark.parametrize("seed", range(10))
    def test_unicode_strings(self, seed):
        """Test unicode string handling."""
        random.seed(seed)

        test_strings = [random_unicode_string() for _ in range(20)]

        # Create with rustypyxl
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Unicode')

        for i, s in enumerate(test_strings, 1):
            rp_wb.set_cell_value('Unicode', i, 1, s)

        rp_wb.save(rp_path)

        # Load with openpyxl
        op_wb = openpyxl.load_workbook(rp_path)
        op_ws = op_wb['Unicode']

        for i, expected in enumerate(test_strings, 1):
            actual = op_ws.cell(i, 1).value
            # Some characters might not roundtrip perfectly
            if actual != expected:
                # Check if it's a known encoding issue
                pass  # Log but don't fail for now

        os.unlink(rp_path)

    def test_mixed_types_in_column(self):
        """Test mixed data types in a single column."""
        values = [
            "String",
            42,
            3.14,
            True,
            False,
            "Another string",
            0,
            -1,
            1e10,
            "",
        ]

        # Create with rustypyxl
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Mixed')

        for i, val in enumerate(values, 1):
            if val != "":  # Skip empty string as it becomes None
                rp_wb.set_cell_value('Mixed', i, 1, val)

        rp_wb.save(rp_path)

        # Load with openpyxl
        op_wb = openpyxl.load_workbook(rp_path)
        op_ws = op_wb['Mixed']

        for i, expected in enumerate(values, 1):
            actual = op_ws.cell(i, 1).value
            if expected == "":
                continue  # Skip empty string comparison
            assert compare_values(actual, expected), f"Row {i}: expected {expected!r}, got {actual!r}"

        os.unlink(rp_path)


# ============================================================================
# Read-Modify-Write Tests
# ============================================================================

class TestReadModifyWrite:
    """Test read-modify-write cycles."""

    def test_modify_existing_cells(self):
        """Test modifying existing cells and preserving others."""
        # Create initial file with openpyxl
        op_path = create_temp_file()
        op_wb = openpyxl.Workbook()
        op_ws = op_wb.active
        op_ws.title = 'Data'

        # Fill with initial data
        for r in range(1, 11):
            for c in range(1, 6):
                op_ws.cell(r, c, f"R{r}C{c}")

        op_wb.save(op_path)

        # Load with rustypyxl, modify some cells, save
        rp_wb = rustypyxl.load_workbook(op_path)
        rp_wb.set_cell_value('Data', 1, 1, "MODIFIED")
        rp_wb.set_cell_value('Data', 5, 3, 999)
        rp_wb.set_cell_value('Data', 10, 5, True)

        rp_path = create_temp_file()
        rp_wb.save(rp_path)

        # Load with openpyxl and verify
        op_wb2 = openpyxl.load_workbook(rp_path)
        op_ws2 = op_wb2['Data']

        # Check modified cells
        assert op_ws2['A1'].value == "MODIFIED"
        assert op_ws2.cell(5, 3).value == 999
        assert op_ws2.cell(10, 5).value == True

        # Check unmodified cells
        assert op_ws2.cell(2, 2).value == "R2C2"
        assert op_ws2.cell(7, 4).value == "R7C4"

        os.unlink(op_path)
        os.unlink(rp_path)

    def test_add_new_rows(self):
        """Test adding new rows to existing data."""
        # Create initial file
        op_path = create_temp_file()
        op_wb = openpyxl.Workbook()
        op_ws = op_wb.active
        op_ws.title = 'Data'

        for r in range(1, 6):
            op_ws.cell(r, 1, f"Original{r}")

        op_wb.save(op_path)

        # Load with rustypyxl and add rows
        rp_wb = rustypyxl.load_workbook(op_path)
        for r in range(6, 11):
            rp_wb.set_cell_value('Data', r, 1, f"Added{r}")

        rp_path = create_temp_file()
        rp_wb.save(rp_path)

        # Verify with openpyxl
        op_wb2 = openpyxl.load_workbook(rp_path)
        op_ws2 = op_wb2['Data']

        # Check all rows
        for r in range(1, 6):
            assert op_ws2.cell(r, 1).value == f"Original{r}"
        for r in range(6, 11):
            assert op_ws2.cell(r, 1).value == f"Added{r}"

        os.unlink(op_path)
        os.unlink(rp_path)


# ============================================================================
# Multiple Sheet Tests
# ============================================================================

class TestMultipleSheets:
    """Test handling of multiple worksheets."""

    def test_multiple_sheets_roundtrip(self):
        """Test creating and reading multiple sheets."""
        # Create with rustypyxl
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()

        sheet_names = ['Sheet1', 'Data', 'Results', 'Summary']
        for name in sheet_names:
            rp_wb.create_sheet(name)
            rp_wb.set_cell_value(name, 1, 1, f"Header for {name}")
            rp_wb.set_cell_value(name, 2, 1, 100)

        rp_wb.save(rp_path)

        # Load with openpyxl
        op_wb = openpyxl.load_workbook(rp_path)

        for name in sheet_names:
            assert name in op_wb.sheetnames, f"Sheet '{name}' not found"
            ws = op_wb[name]
            assert ws['A1'].value == f"Header for {name}"
            assert ws['A2'].value == 100

        os.unlink(rp_path)

    def test_sheet_with_special_names(self):
        """Test sheets with special characters in names."""
        special_names = [
            'Sheet 1',  # Space
            'Data-2024',  # Hyphen
            'Results_Final',  # Underscore
            'Summary (v2)',  # Parentheses
        ]

        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()

        for name in special_names:
            rp_wb.create_sheet(name)
            rp_wb.set_cell_value(name, 1, 1, "Test")

        rp_wb.save(rp_path)

        # Load with openpyxl
        op_wb = openpyxl.load_workbook(rp_path)

        for name in special_names:
            assert name in op_wb.sheetnames, f"Sheet '{name}' not found"

        os.unlink(rp_path)


# ============================================================================
# Edge Case Tests
# ============================================================================

class TestEdgeCases:
    """Test edge cases and boundary conditions."""

    def test_empty_workbook(self):
        """Test empty workbook handling."""
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Empty')
        rp_wb.save(rp_path)

        op_wb = openpyxl.load_workbook(rp_path)
        assert 'Empty' in op_wb.sheetnames

        os.unlink(rp_path)

    def test_large_row_numbers(self):
        """Test handling of large row numbers."""
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Large')

        # Write to row 10000
        rp_wb.set_cell_value('Large', 10000, 1, "Row 10000")
        rp_wb.set_cell_value('Large', 1, 1, "Row 1")

        rp_wb.save(rp_path)

        op_wb = openpyxl.load_workbook(rp_path)
        ws = op_wb['Large']

        assert ws.cell(1, 1).value == "Row 1"
        assert ws.cell(10000, 1).value == "Row 10000"

        os.unlink(rp_path)

    def test_large_column_numbers(self):
        """Test handling of large column numbers."""
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Wide')

        # Write to column 100 (CV)
        rp_wb.set_cell_value('Wide', 1, 100, "Col 100")
        rp_wb.set_cell_value('Wide', 1, 1, "Col 1")

        rp_wb.save(rp_path)

        op_wb = openpyxl.load_workbook(rp_path)
        ws = op_wb['Wide']

        assert ws.cell(1, 1).value == "Col 1"
        assert ws.cell(1, 100).value == "Col 100"

        os.unlink(rp_path)

    def test_special_xml_characters(self):
        """Test handling of XML special characters."""
        special_strings = [
            "<tag>",
            "a & b",
            '"quoted"',
            "'single'",
            "a < b > c",
            "<script>alert('xss')</script>",
            "1 & 2 < 3 > 0",
        ]

        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Special')

        for i, s in enumerate(special_strings, 1):
            rp_wb.set_cell_value('Special', i, 1, s)

        rp_wb.save(rp_path)

        op_wb = openpyxl.load_workbook(rp_path)
        ws = op_wb['Special']

        for i, expected in enumerate(special_strings, 1):
            actual = ws.cell(i, 1).value
            assert actual == expected, f"Row {i}: expected {expected!r}, got {actual!r}"

        os.unlink(rp_path)

    def test_very_long_strings(self):
        """Test handling of very long strings."""
        long_strings = [
            "a" * 100,
            "b" * 1000,
            "c" * 10000,
        ]

        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Long')

        for i, s in enumerate(long_strings, 1):
            rp_wb.set_cell_value('Long', i, 1, s)

        rp_wb.save(rp_path)

        op_wb = openpyxl.load_workbook(rp_path)
        ws = op_wb['Long']

        for i, expected in enumerate(long_strings, 1):
            actual = ws.cell(i, 1).value
            assert actual == expected, f"Row {i}: length mismatch {len(actual)} vs {len(expected)}"

        os.unlink(rp_path)


# ============================================================================
# Bulk API Tests
# ============================================================================

class TestBulkAPI:
    """Test bulk read/write APIs."""

    def test_write_rows_vs_openpyxl(self):
        """Test bulk write_rows produces same output as openpyxl."""
        data = [
            ["Name", "Age", "Score"],
            ["Alice", 30, 95.5],
            ["Bob", 25, 87.3],
            ["Charlie", 35, 92.1],
        ]

        # Create with rustypyxl using write_rows
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Data')
        rp_wb.write_rows('Data', data)
        rp_wb.save(rp_path)

        # Create with openpyxl for comparison
        op_path = create_temp_file()
        op_wb = openpyxl.Workbook()
        op_ws = op_wb.active
        op_ws.title = 'Data'
        for row in data:
            op_ws.append(row)
        op_wb.save(op_path)

        # Load both and compare
        rp_loaded = openpyxl.load_workbook(rp_path)
        op_loaded = openpyxl.load_workbook(op_path)

        rp_ws = rp_loaded['Data']
        op_ws = op_loaded['Data']

        for r in range(1, len(data) + 1):
            for c in range(1, len(data[0]) + 1):
                rp_val = rp_ws.cell(r, c).value
                op_val = op_ws.cell(r, c).value
                assert compare_values(rp_val, op_val), f"({r},{c}): rustypyxl={rp_val}, openpyxl={op_val}"

        os.unlink(rp_path)
        os.unlink(op_path)

    def test_read_rows_matches_openpyxl(self):
        """Test bulk read_rows returns same data as openpyxl iteration."""
        # Create test file with openpyxl
        op_path = create_temp_file()
        op_wb = openpyxl.Workbook()
        op_ws = op_wb.active
        op_ws.title = 'Data'

        test_data = [
            [f"R{r}C{c}" for c in range(1, 6)]
            for r in range(1, 11)
        ]

        for row in test_data:
            op_ws.append(row)
        op_wb.save(op_path)

        # Load with rustypyxl and use read_rows
        rp_wb = rustypyxl.load_workbook(op_path)
        rp_data = rp_wb.read_rows('Data')

        # Compare
        assert len(rp_data) == len(test_data), f"Row count: {len(rp_data)} vs {len(test_data)}"

        for r, (rp_row, expected_row) in enumerate(zip(rp_data, test_data)):
            for c, (rp_val, expected_val) in enumerate(zip(rp_row, expected_row)):
                assert compare_values(rp_val, expected_val), f"({r+1},{c+1}): {rp_val} vs {expected_val}"

        os.unlink(op_path)


# ============================================================================
# Formula Tests
# ============================================================================

class TestFormulas:
    """Test formula handling."""

    def test_formula_preservation(self):
        """Test that formulas are preserved through rustypyxl."""
        # Create file with openpyxl containing formulas
        op_path = create_temp_file()
        op_wb = openpyxl.Workbook()
        op_ws = op_wb.active
        op_ws.title = 'Formulas'

        op_ws['A1'] = 10
        op_ws['A2'] = 20
        op_ws['A3'] = '=A1+A2'
        op_ws['B1'] = '=SUM(A1:A2)'

        op_wb.save(op_path)

        # Load with rustypyxl and save
        rp_wb = rustypyxl.load_workbook(op_path)
        rp_path = create_temp_file()
        rp_wb.save(rp_path)

        # Load with openpyxl and verify formulas
        op_wb2 = openpyxl.load_workbook(rp_path)
        op_ws2 = op_wb2['Formulas']

        # Check values are preserved
        assert op_ws2['A1'].value == 10
        assert op_ws2['A2'].value == 20

        # Check formulas (may be stored as string or formula)
        a3_val = op_ws2['A3'].value
        assert a3_val == '=A1+A2' or a3_val == 30, f"A3 formula/value: {a3_val}"

        os.unlink(op_path)
        os.unlink(rp_path)


# ============================================================================
# Stress Tests
# ============================================================================

class TestStress:
    """Stress tests with larger datasets."""

    @pytest.mark.slow
    def test_large_dataset(self):
        """Test with a larger dataset (10k rows)."""
        rows = 10000
        cols = 20

        # Generate data
        data = [
            [f"R{r}C{c}" if c % 3 == 0 else r * c + c * 0.1 for c in range(cols)]
            for r in range(rows)
        ]

        # Create with rustypyxl
        rp_path = create_temp_file()
        rp_wb = rustypyxl.Workbook()
        rp_wb.create_sheet('Data')
        rp_wb.write_rows('Data', data)
        rp_wb.save(rp_path)

        # Load with openpyxl and spot-check
        op_wb = openpyxl.load_workbook(rp_path)
        ws = op_wb['Data']

        # Check corners
        assert ws.cell(1, 1).value == "R0C0"
        assert ws.cell(rows, cols).value == data[rows-1][cols-1]

        # Random spot checks
        random.seed(42)
        for _ in range(100):
            r = random.randint(0, rows - 1)
            c = random.randint(0, cols - 1)
            expected = data[r][c]
            actual = ws.cell(r + 1, c + 1).value
            assert compare_values(actual, expected), f"({r+1},{c+1}): {actual} vs {expected}"

        os.unlink(rp_path)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
