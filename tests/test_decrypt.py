"""Reading a password-protected (agile-encrypted) workbook from Python. The
fixture is produced at test time by msoffcrypto-tool (an independent
implementation), so a successful read cross-validates the decryption.

openpyxl cannot open encrypted files at all, so this is a rustypyxl
differentiator, exposed as load_workbook(source, password=...).
"""

import io

import msoffcrypto
import openpyxl
import pytest
import rustypyxl
from msoffcrypto.format.ooxml import OOXMLFile


def _encrypted_bytes(password="s3cret"):
    # A plain workbook with known content, encrypted with msoffcrypto.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Secret"
    ws["A1"] = "hello"
    ws["B1"] = 42
    plain = io.BytesIO()
    wb.save(plain)
    plain.seek(0)
    enc = io.BytesIO()
    OOXMLFile(plain).encrypt(password, enc)
    return enc.getvalue()


def test_load_encrypted_with_password(tmp_path):
    enc = _encrypted_bytes()
    path = tmp_path / "protected.xlsx"
    path.write_bytes(enc)

    wb = rustypyxl.load_workbook(str(path), password="s3cret")
    ws = wb["Secret"]
    assert ws["A1"].value == "hello"
    assert ws["B1"].value == 42


def test_load_encrypted_from_bytes():
    wb = rustypyxl.load_workbook(_encrypted_bytes(), password="s3cret")
    assert wb["Secret"]["A1"].value == "hello"


def test_wrong_password_raises():
    with pytest.raises(ValueError):
        rustypyxl.load_workbook(_encrypted_bytes(), password="wrong")


def test_encrypted_without_password_gives_clear_error():
    with pytest.raises(ValueError, match="encrypted"):
        rustypyxl.load_workbook(_encrypted_bytes())
