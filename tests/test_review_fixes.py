"""Regression tests for the consistency/usability review fixes."""

import io
import os
import tempfile

import openpyxl
import pytest
import rustypyxl


def reopen(wb, sheet_name):
    data = wb.save_to_bytes()
    return openpyxl.load_workbook(io.BytesIO(data))[sheet_name]


class TestUnderlineFidelity:
    """Underline style is preserved, not collapsed to a bool."""

    def test_double_underline_round_trips(self, workbook_with_sheet):
        wb = workbook_with_sheet
        ws = wb.active
        ws["A1"] = "x"
        ws["A1"].font = rustypyxl.Font(underline="double")
        op = reopen(wb, "Test")
        assert op["A1"].font.underline == "double"

    def test_single_underline_round_trips(self, workbook_with_sheet):
        wb = workbook_with_sheet
        ws = wb.active
        ws["A1"] = "x"
        ws["A1"].font = rustypyxl.Font(underline="single")
        op = reopen(wb, "Test")
        assert op["A1"].font.underline == "single"


class TestCellMetadataPersistence:
    """Hyperlink and comment setters write through to the workbook."""

    def test_hyperlink_persists_to_workbook(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws["A1"].hyperlink = "https://example.com"
        # A fresh handle reads it back from the workbook, not transient cell state.
        assert ws["A1"].hyperlink == "https://example.com"

    def test_comment_persists_to_workbook(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws["A1"].comment = "a note"
        assert ws["A1"].comment == "a note"


class TestStyleClassRegistration:
    def test_color_is_exposed(self):
        c = rustypyxl.Color(rgb="FFFF0000")
        assert c.rgb == "FFFF0000"
        assert rustypyxl.styles.Color is rustypyxl.Color


class TestCreateSheetIndex:
    def test_create_sheet_honors_index(self, empty_workbook):
        wb = empty_workbook
        wb.create_sheet("A")
        wb.create_sheet("B")
        wb.create_sheet("First", index=0)
        assert wb.sheetnames == ["First", "A", "B"]

    def test_create_sheet_appends_by_default(self, empty_workbook):
        wb = empty_workbook
        wb.create_sheet("A")
        wb.create_sheet("B")
        assert wb.sheetnames == ["A", "B"]


class TestStreamingFormula:
    """WriteOnlyWorkbook stores formulas consistently with the normal path."""

    def test_streaming_formula_has_single_equals(self):
        path = tempfile.mktemp(suffix=".xlsx")
        try:
            sw = rustypyxl.WriteOnlyWorkbook(path)
            sw.create_sheet("S")
            sw.append_row([1, 2, "=SUM(A1:B1)"])
            sw.close()
            op = openpyxl.load_workbook(path)["S"]
            assert op["C1"].value == "=SUM(A1:B1)"
        finally:
            if os.path.exists(path):
                os.unlink(path)
