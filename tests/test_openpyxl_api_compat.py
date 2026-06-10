"""Regression tests for openpyxl API compatibility (task 009)."""

import gc
import pathlib

import pytest

import rustypyxl
from rustypyxl import Color, Font, PatternFill, Side


class TestLazyIteration:
    def test_iter_rows_returns_iterator_of_tuples(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append([1, "a"])
        ws.append([2, "b"])
        it = ws.iter_rows(values_only=True)
        assert next(it) == (1, "a")
        assert next(it) == (2, "b")
        with pytest.raises(StopIteration):
            next(it)

    def test_iter_rows_cells_are_tuples(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append(["x"])
        row = next(ws.iter_rows())
        assert isinstance(row, tuple)
        assert row[0].value == "x"

    def test_iterator_raises_on_removed_sheet(self):
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("A")
        wb.create_sheet("B")
        ws.append([1])
        it = ws.iter_rows(values_only=True)
        wb.remove(wb["A"])
        with pytest.raises(ValueError):
            next(it)


class TestStyleConstructorCompat:
    def test_patternfill_start_end_color(self):
        fill = PatternFill(start_color="FFFF00", end_color="FF0000", fill_type="solid")
        assert fill.fgColor == "FFFF00"
        assert fill.bgColor == "FF0000"
        assert fill.start_color == "FFFF00"
        assert fill.end_color == "FF0000"

    def test_patternfill_applies_to_cell(self, workbook_with_sheet, temp_xlsx_path):
        import openpyxl

        ws = workbook_with_sheet.active
        ws["A1"] = "filled"
        ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        workbook_with_sheet.save(temp_xlsx_path)

        chk = openpyxl.load_workbook(temp_xlsx_path)
        assert chk["Test"]["A1"].fill.fgColor.rgb.endswith("FFFF00")

    def test_font_accepts_color_object(self):
        font = Font(color=Color(rgb="FF0000"), bold=True)
        assert font.color == "FF0000"

    def test_side_accepts_color_object(self):
        side = Side(style="thin", color=Color(rgb="00FF00"))
        assert side.color == "00FF00"

    def test_color_string_still_accepted(self):
        assert Font(color="123456").color == "123456"


class TestValueTypes:
    def test_integral_numbers_return_int(self, workbook_with_sheet, temp_xlsx_path):
        ws = workbook_with_sheet.active
        ws["A1"] = 42
        ws["A2"] = 1.5
        workbook_with_sheet.save(temp_xlsx_path)

        wb = rustypyxl.load_workbook(temp_xlsx_path)
        ws2 = wb["Test"]
        assert ws2["A1"].value == 42
        assert isinstance(ws2["A1"].value, int), "integral cell should read as int"
        assert isinstance(ws2["A2"].value, float)


class TestAppendInputs:
    def test_append_generator(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append(x * 2 for x in range(3))
        assert ws["A1"].value == 0
        assert ws["C1"].value == 4

    def test_append_tuple(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append((1, 2))
        assert ws["B1"].value == 2

    def test_append_dict_with_column_letters(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append({"A": "first", "C": "third"})
        assert ws["A1"].value == "first"
        assert ws["B1"].value is None
        assert ws["C1"].value == "third"

    def test_append_dict_with_indices(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append({1: "a", 3: "c"})
        assert ws["A1"].value == "a"
        assert ws["C1"].value == "c"

    def test_append_dict_bad_key(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        with pytest.raises(ValueError):
            ws.append({"not a column": 1})


class TestGarbageCollection:
    def test_cell_reference_cycle_is_collectable(self):
        cell = rustypyxl.Cell(1, 1)
        cell.value = cell  # self-cycle through value_internal
        del cell
        assert gc.collect() >= 1, "cycle through Cell was not collected"

    def test_worksheet_handle_cycle_is_collectable(self):
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("A")
        holder = {"ws": ws, "wb": wb}
        holder["self"] = holder
        del wb, ws, holder
        gc.collect()  # must not crash; handles participate in GC


class TestTypeStubs:
    def test_stubs_shipped_with_package(self):
        pkg_dir = pathlib.Path(rustypyxl.__file__).parent
        assert (pkg_dir / "__init__.pyi").exists(), "type stubs missing from wheel"
        assert (pkg_dir / "py.typed").exists(), "py.typed marker missing"
