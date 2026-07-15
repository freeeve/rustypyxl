"""Sheet protection set through the Python binding, read back by openpyxl."""

import openpyxl
import rustypyxl


def test_protect_and_unprotect(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws["A1"] = "x"
    assert ws.sheet_protected is False

    ws.protect_sheet(password="secret")
    assert ws.sheet_protected is True

    out = str(tmp_path / "prot.xlsx")
    wb.save(out)
    ows = openpyxl.load_workbook(out)["S"]
    assert ows.protection.sheet is True

    # Unprotecting clears it.
    ws.unprotect_sheet()
    assert ws.sheet_protected is False


def test_protect_without_password(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws["A1"] = "x"
    ws.protect_sheet()
    out = str(tmp_path / "prot2.xlsx")
    wb.save(out)
    assert openpyxl.load_workbook(out)["S"].protection.sheet is True
