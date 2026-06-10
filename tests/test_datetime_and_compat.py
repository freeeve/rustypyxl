"""Tests for datetime cell support and openpyxl-compat fixes."""

import datetime
import threading
import time

import openpyxl
import pytest

import rustypyxl


class TestDatetimeValues:
    def test_datetime_roundtrip(self, temp_xlsx_path):
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        dt = datetime.datetime(2024, 3, 15, 10, 30, 45)
        ws["A1"] = dt
        wb.save(temp_xlsx_path)

        wb2 = rustypyxl.load_workbook(temp_xlsx_path)
        assert wb2["Data"]["A1"].value == dt

    def test_date_roundtrip(self, temp_xlsx_path):
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        d = datetime.date(2024, 3, 15)
        ws["A1"] = d
        wb.save(temp_xlsx_path)

        wb2 = rustypyxl.load_workbook(temp_xlsx_path)
        assert wb2["Data"]["A1"].value == d

    def test_openpyxl_reads_datetime_cells(self, temp_xlsx_path):
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        dt = datetime.datetime(2024, 3, 15, 10, 30, 45)
        d = datetime.date(2024, 3, 15)
        ws["A1"] = dt
        ws["A2"] = d
        wb.save(temp_xlsx_path)

        owb = openpyxl.load_workbook(temp_xlsx_path)
        ows = owb["Data"]
        assert ows["A1"].value == dt
        assert ows["A2"].value == d

    def test_streaming_datetime_not_dropped(self, temp_xlsx_path):
        dt = datetime.datetime(2025, 1, 2, 3, 4, 5)
        swb = rustypyxl.WriteOnlyWorkbook(str(temp_xlsx_path))
        swb.create_sheet("S")
        swb.append_row([dt, 7])
        swb.close()

        wb = rustypyxl.load_workbook(temp_xlsx_path)
        assert wb["S"]["A1"].value == dt
        assert wb["S"]["B1"].value == 7


class TestNonFiniteFloats:
    def test_nan_and_inf_produce_loadable_file(self, temp_xlsx_path):
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        ws["A1"] = float("nan")
        ws["A2"] = float("inf")
        ws["A3"] = float("-inf")
        wb.save(temp_xlsx_path)

        # The file must remain valid for both readers.
        rustypyxl.load_workbook(temp_xlsx_path)
        openpyxl.load_workbook(temp_xlsx_path)


class TestOpenpyxlCompatErrors:
    def test_missing_sheet_raises_keyerror(self):
        wb = rustypyxl.Workbook()
        wb.create_sheet("Only")
        with pytest.raises(KeyError):
            _ = wb["NonExistent"]

    def test_load_and_save_accept_pathlib_path(self, tmp_path):
        path = tmp_path / "pathlib_test.xlsx"  # a pathlib.Path, not str
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        ws["A1"] = "hello"
        wb.save(path)

        wb2 = rustypyxl.load_workbook(path)
        assert wb2["Data"]["A1"].value == "hello"

    def test_styles_submodule_importable(self):
        from rustypyxl.styles import Font

        assert Font is not None

    def test_streaming_early_close_does_not_brick_workbook(self, temp_xlsx_path):
        swb = rustypyxl.WriteOnlyWorkbook(str(temp_xlsx_path))
        with pytest.raises(ValueError):
            swb.close()
        # The workbook must still be usable after the failed close.
        swb.create_sheet("S")
        swb.append_row(["ok"])
        swb.close()

        wb = rustypyxl.load_workbook(temp_xlsx_path)
        assert wb["S"]["A1"].value == "ok"


class TestGilRelease:
    def test_save_to_bytes_releases_gil(self):
        wb = rustypyxl.Workbook()
        wb.create_sheet("B")
        wb.write_rows("B", [[i, "x" * 20] for i in range(200_000)])

        ticks = []
        stop = threading.Event()

        def ticker():
            while not stop.is_set():
                ticks.append(1)
                time.sleep(0.002)

        t = threading.Thread(target=ticker)
        t.start()
        try:
            wb.save_to_bytes()
        finally:
            stop.set()
            t.join()
        # If the GIL were held for the whole save, the ticker could not run.
        assert len(ticks) > 0
