"""Tests for data validation support."""

import openpyxl

import rustypyxl


class TestValidationRoundtrip:
    """Test data validation survives save/load cycle."""

    def test_load_existing_validation(self, fixtures_dir):
        """Load an externally-authored file with data validation."""
        wb = rustypyxl.load_workbook(str(fixtures_dir / "validation.xlsx"))
        assert wb.sheetnames == ["Validated"]
        assert wb["Validated"]["A1"].value == "pick one"

    def test_validation_survives_roundtrip(self, fixtures_dir, temp_xlsx_path):
        """The validation rule must survive load+save."""
        wb = rustypyxl.load_workbook(str(fixtures_dir / "validation.xlsx"))
        wb.save(temp_xlsx_path)

        chk = openpyxl.load_workbook(temp_xlsx_path)
        rules = chk["Validated"].data_validations.dataValidation
        assert len(rules) == 1, "validation rule stripped on round-trip"
        assert rules[0].type == "list"
        assert rules[0].formula1 == '"Yes,No,Maybe"'
        assert str(rules[0].sqref) == "B1:B10"


class TestValidationMessages:
    """The error/prompt dialog text must survive a load->save cycle."""

    def _authored(self, path):
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "V"
        dv = DataValidation(
            type="whole",
            operator="greaterThan",
            formula1="10",
            allow_blank=False,
            showErrorMessage=True,
            showInputMessage=True,
        )
        dv.errorStyle = "warning"
        dv.errorTitle = "Too small"
        dv.error = "Enter a number above 10."
        dv.promptTitle = "Quantity"
        dv.prompt = "How many units?"
        dv.add("A1:A20")
        ws.add_data_validation(dv)
        wb.save(path)

    def test_error_and_prompt_text_survive_roundtrip(self, tmp_path, temp_xlsx_path):
        import openpyxl

        src = str(tmp_path / "authored.xlsx")
        self._authored(src)

        wb = rustypyxl.load_workbook(src)
        wb.save(temp_xlsx_path)

        rule = openpyxl.load_workbook(temp_xlsx_path)["V"].data_validations.dataValidation[0]
        assert rule.type == "whole"
        assert rule.operator == "greaterThan"
        assert rule.formula1 == "10"
        assert rule.errorStyle == "warning"
        assert rule.errorTitle == "Too small"
        assert rule.error == "Enter a number above 10."
        assert rule.promptTitle == "Quantity"
        assert rule.prompt == "How many units?"
