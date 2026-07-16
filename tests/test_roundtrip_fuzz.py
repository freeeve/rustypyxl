"""Fuzz the Python binding's save -> load round-trip cell-by-cell.

A seeded RNG (deterministic, so failures reproduce) fills a grid with random
values of every supported type, saves, reloads, and checks each cell. Two
intentional normalizations are accounted for:
  * an integral number reads back as int (openpyxl does the same), and
  * a cell set to None / "" is an empty cell on reload.
"""

import random

import openpyxl
import pytest
import rustypyxl

# A charset without control characters (those are illegal in XML and stripped).
_CHARS = "abcdefgABCDEFG0123 _-.,:/😀✓é—\t"


def _rand_value(rng):
    kind = rng.randrange(5)
    if kind == 0:
        return "".join(rng.choice(_CHARS) for _ in range(rng.randrange(0, 30)))
    if kind == 1:
        return rng.randint(-10**12, 10**12)
    if kind == 2:
        # non-integral float (integral floats normalize to int; see _expect)
        return rng.uniform(-1e6, 1e6) + rng.choice([0.123, 0.5, 0.777])
    if kind == 3:
        return rng.choice([True, False])
    return "=" + rng.choice(["SUM(A1:A2)", "A1*2", "1+1"])


def _expect(v):
    """The value we expect back after a round-trip, applying the documented
    normalizations."""
    if isinstance(v, bool):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


@pytest.mark.parametrize("seed", range(6))
def test_binding_roundtrip_fuzz(seed):
    rng = random.Random(seed)
    ws_name = "S"
    cells = {}
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet(ws_name)
    for _ in range(400):
        row = rng.randint(1, 400)
        col = rng.randint(1, 30)
        v = _rand_value(rng)
        cells[(row, col)] = v
        ws.cell(row=row, column=col).value = v

    reloaded = rustypyxl.load_workbook(wb.save_to_bytes())[ws_name]
    for (row, col), sent in cells.items():
        got = reloaded.cell(row=row, column=col).value
        exp = _expect(sent)
        if isinstance(exp, float):
            assert got == pytest.approx(exp, rel=1e-12, abs=1e-9), f"({row},{col})"
        elif isinstance(exp, str) and exp.startswith("="):
            # formulas read back as their "=..." text
            assert got == exp, f"({row},{col})"
        elif exp == "":
            assert got in (None, ""), f"({row},{col})"
        else:
            assert got == exp, f"({row},{col}) sent {sent!r} got {got!r}"


def test_openpyxl_empty_string_is_preserved(tmp_path):
    # Known interop nuance: openpyxl writes "" as an empty inline/shared string.
    # Confirm what rustypyxl does with it on load (documents the behavior).
    src = str(tmp_path / "empty.xlsx")
    wb = openpyxl.Workbook()
    wb.active["A1"] = ""
    wb.active["A2"] = "x"
    wb.save(src)

    ws = rustypyxl.load_workbook(src).active
    # A2 (a real value) must survive regardless.
    assert ws["A2"].value == "x"
    # A1 ("") comes back as None or "" -- both are acceptable renderings of an
    # empty cell; assert it does not become something else.
    assert ws["A1"].value in (None, "")
