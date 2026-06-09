"""Tests for Worksheet functionality.

These exercise the openpyxl-style Worksheet API end to end: where a behavior
must persist, the workbook is saved to bytes and reloaded with openpyxl to
confirm the output is real (not a no-op).
"""

import io

import openpyxl
import pytest
import rustypyxl


def reload_with_openpyxl(wb, sheet_name):
    """Save a rustypyxl workbook to bytes and reopen the sheet with openpyxl."""
    data = wb.save_to_bytes()
    return openpyxl.load_workbook(io.BytesIO(data))[sheet_name]


class TestWorksheetBasics:
    """Test basic worksheet operations."""

    def test_worksheet_title(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        assert ws.title == "Test"

    def test_worksheet_repr(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        assert "Worksheet" in str(ws)
        assert "Test" in str(ws)

    def test_rename_persists(self, workbook_with_sheet):
        """Renaming a sheet updates the workbook and the saved file."""
        wb = workbook_with_sheet
        ws = wb.active
        ws.title = "Results"
        assert ws.title == "Results"
        assert wb.sheetnames == ["Results"]
        op = reload_with_openpyxl(wb, "Results")
        assert op.title == "Results"

    def test_rename_to_existing_raises(self, empty_workbook):
        wb = empty_workbook
        wb.create_sheet("One")
        wb.create_sheet("Two")
        ws_two = wb["Two"]
        with pytest.raises(ValueError):
            ws_two.title = "One"


class TestCellAccess:
    """Test cell access methods."""

    def test_cell_by_coordinate(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        cell = ws["A1"]
        assert cell.coordinate == "A1"

    def test_cell_by_row_col(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        cell = ws.cell(1, 1)
        assert cell.row == 1
        assert cell.column == 1

    def test_cell_invalid_row_raises(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        with pytest.raises(ValueError):
            ws.cell(0, 1)

    def test_cell_invalid_column_raises(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        with pytest.raises(ValueError):
            ws.cell(1, 0)

    def test_setitem_persists(self, workbook_with_sheet):
        """ws['A1'] = value writes through to the workbook and the file."""
        wb = workbook_with_sheet
        ws = wb.active
        ws["A1"] = "Hello"
        ws["B2"] = 42
        assert ws["A1"].value == "Hello"
        assert ws["B2"].value == 42
        op = reload_with_openpyxl(wb, "Test")
        assert op["A1"].value == "Hello"
        assert op["B2"].value == 42

    def test_cell_value_setter_persists(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.cell(3, 1).value = "world"
        assert ws.cell(3, 1).value == "world"

    def test_setitem_range_raises(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        with pytest.raises(ValueError):
            ws["A1:B2"] = 1

    def test_getitem_range_returns_cells(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws["A1"] = 1
        ws["B2"] = 2
        block = ws["A1:B2"]
        assert len(block) == 2  # two rows
        assert len(block[0]) == 2  # two columns
        assert block[0][0].coordinate == "A1"
        assert block[1][1].coordinate == "B2"

    def test_data_type(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws["A1"] = "text"
        ws["A2"] = 3.5
        ws["A3"] = True
        ws["A4"] = "=SUM(A2:A2)"
        assert ws["A1"].data_type == "s"
        assert ws["A2"].data_type == "n"
        assert ws["A3"].data_type == "b"
        assert ws["A4"].data_type == "f"

    def test_offset_stays_connected(self, workbook_with_sheet):
        """offset() returns a cell that still writes back to the workbook."""
        ws = workbook_with_sheet.active
        ws["A1"].offset(0, 1).value = "B1val"
        assert ws["B1"].value == "B1val"


class TestIterRows:
    """Iteration returns real data, not fabricated empties."""

    def test_iter_rows_values(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])
        rows = ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True)
        assert rows == [[1, 2, 3], [4, 5, 6]]

    def test_iter_rows_cells(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append(["a", "b"])
        rows = ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2)
        assert rows[0][0].value == "a"
        assert rows[0][1].value == "b"

    def test_iter_rows_defaults_to_used_range(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append([1, 2])
        ws.append([3, 4])
        rows = ws.iter_rows(values_only=True)
        assert rows == [[1, 2], [3, 4]]


class TestIterCols:
    def test_iter_cols_values(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws.append([1, 2])
        ws.append([3, 4])
        cols = ws.iter_cols(min_col=1, max_col=2, min_row=1, max_row=2, values_only=True)
        assert cols == [[1, 3], [2, 4]]


class TestDimensions:
    """Dimensions reflect the actual extent of the data."""

    def test_dimensions_reflect_data(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        ws["B2"] = "x"
        ws["D5"] = "y"
        assert ws.max_row == 5
        assert ws.max_column == 4
        assert ws.min_row == 2
        assert ws.min_column == 2
        assert ws.dimensions == "B2:D5"

    def test_dimensions_empty_sheet(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        assert ws.dimensions == "A1:A1"


class TestAppend:
    def test_append_positions(self, workbook_with_sheet):
        wb = workbook_with_sheet
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == 30
        op = reload_with_openpyxl(wb, "Test")
        assert op["A1"].value == "Name"
        assert op["B2"].value == 30


class TestFreezePanes:
    def test_freeze_panes_round_trip(self, workbook_with_sheet):
        wb = workbook_with_sheet
        ws = wb.active
        ws["A1"] = "data"
        ws.freeze_panes = "B2"
        assert ws.freeze_panes == "B2"
        op = reload_with_openpyxl(wb, "Test")
        assert op.freeze_panes == "B2"


class TestRowColumnOperations:
    """Structural row/column edits are not implemented yet and say so."""

    def test_insert_rows_not_implemented(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        with pytest.raises(NotImplementedError):
            ws.insert_rows(1, 2)

    def test_insert_cols_not_implemented(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        with pytest.raises(NotImplementedError):
            ws.insert_cols(1, 2)

    def test_delete_rows_not_implemented(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        with pytest.raises(NotImplementedError):
            ws.delete_rows(1, 2)

    def test_delete_cols_not_implemented(self, workbook_with_sheet):
        ws = workbook_with_sheet.active
        with pytest.raises(NotImplementedError):
            ws.delete_cols(1, 2)
