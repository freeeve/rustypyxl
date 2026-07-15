"""AutoFilter set through the openpyxl-style ws.auto_filter.ref idiom, read back
by openpyxl.
"""

import openpyxl
import rustypyxl


def test_auto_filter_ref_round_trips(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    wb.write_rows("S", [["Name", "Qty"], ["Widget", 3]])
    assert ws.auto_filter.ref is None

    ws.auto_filter.ref = "A1:B2"
    assert ws.auto_filter.ref == "A1:B2"

    out = str(tmp_path / "af.xlsx")
    wb.save(out)
    ows = openpyxl.load_workbook(out)["S"]
    assert ows.auto_filter.ref == "A1:B2"
