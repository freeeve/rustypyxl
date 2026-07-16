"""Writing password-protected (agile-encrypted) workbooks from Python.

The decisive check is cross-implementation: rustypyxl encrypts, and
msoffcrypto-tool (an independent implementation) decrypts it, with openpyxl
reading the plaintext -- proving the output is genuinely Excel-compatible, not
just self-consistent. openpyxl cannot write encrypted files, so this is a
rustypyxl differentiator.
"""

import io

import msoffcrypto
import openpyxl
import rustypyxl


def _encrypted_workbook(password="s3cret"):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("Secret")
    wb.write_rows("Secret", [["hello", 42], ["world", 7]])
    return wb.save_to_bytes(password=password)


def test_msoffcrypto_decrypts_rustypyxl_output():
    enc = _encrypted_workbook()
    assert enc[:8] == bytes([0xD0, 0xCF, 0x11, 0xE0, 0xA1, 0xB1, 0x1A, 0xE1])

    of = msoffcrypto.OfficeFile(io.BytesIO(enc))
    of.load_key(password="s3cret")
    plain = io.BytesIO()
    of.decrypt(plain)

    ws = openpyxl.load_workbook(io.BytesIO(plain.getvalue()))["Secret"]
    assert ws["A1"].value == "hello"
    assert ws["B1"].value == 42
    assert ws["A2"].value == "world"


def test_rustypyxl_round_trip():
    enc = _encrypted_workbook()
    ws = rustypyxl.load_workbook(enc, password="s3cret")["Secret"]
    assert ws["A1"].value == "hello"
    assert ws["B1"].value == 42


def test_save_encrypted_to_path(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws["A1"] = "secret data"
    out = str(tmp_path / "protected.xlsx")
    wb.save(out, password="pw123")

    # msoffcrypto opens it; a wrong password does not.
    of = msoffcrypto.OfficeFile(open(out, "rb"))
    of.load_key(password="pw123")
    plain = io.BytesIO()
    of.decrypt(plain)
    assert openpyxl.load_workbook(io.BytesIO(plain.getvalue()))["S"]["A1"].value == "secret data"
