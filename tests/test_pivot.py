"""Pivot tables: rustypyxl reads a loaded pivot's definition, creates a pivot
from a source range, and the created file is readable by openpyxl.
"""

import openpyxl
import rustypyxl


def _source_workbook():
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("Sales")
    wb.write_rows(
        "Sales",
        [
            ["Region", "Product", "Amount"],
            ["East", "Widget", 100],
            ["West", "Widget", 150],
            ["East", "Gadget", 200],
        ],
    )
    return wb


def test_create_pivot_reads_back_through_rustypyxl(tmp_path):
    wb = _source_workbook()
    wb.add_pivot_table(
        "Sales",
        "A1:C4",
        "Sales",
        "F1",
        rows=["Region"],
        values=[("Amount", "sum")],
        name="ByRegion",
    )
    out = str(tmp_path / "pivot.xlsx")
    wb.save(out)

    pivots = rustypyxl.load_workbook(out).pivot_tables
    assert len(pivots) == 1
    p = pivots[0]
    assert p.name == "ByRegion"
    assert p.source_sheet == "Sales"
    assert p.source_ref == "A1:C4"
    assert p.fields == ["Region", "Product", "Amount"]
    assert p.row_fields == ["Region"]
    assert p.data_fields == [
        {"name": "Sum of Amount", "source_field": "Amount", "subtotal": "sum"}
    ]


def test_created_pivot_opens_in_openpyxl(tmp_path):
    wb = _source_workbook()
    wb.add_pivot_table(
        "Sales",
        "A1:C4",
        "Sales",
        "F1",
        rows=["Region"],
        columns=["Product"],
        values=[("Amount", "sum")],
    )
    out = str(tmp_path / "pivot.xlsx")
    wb.save(out)

    # openpyxl parses the workbook and finds the pivot table without error.
    loaded = openpyxl.load_workbook(out)
    ws = loaded["Sales"]
    assert len(ws._pivots) == 1
    assert ws._pivots[0].cache is not None


def test_unknown_field_raises(tmp_path):
    wb = _source_workbook()
    try:
        wb.add_pivot_table(
            "Sales", "A1:C4", "Sales", "F1", rows=["Nope"], values=[("Amount", "sum")]
        )
        assert False, "expected ValueError"
    except ValueError:
        pass
