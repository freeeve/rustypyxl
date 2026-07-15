"""Conditional-formatting rules added through the Python binding, read back by
openpyxl.
"""

import openpyxl
import rustypyxl


def _rules(path, sheet="S"):
    ws = openpyxl.load_workbook(path)[sheet]
    out = []
    for rng in ws.conditional_formatting:
        out.extend(rng.rules)
    return out


def test_cell_is_rule(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws.add_conditional_formatting(
        "A1:A10",
        {"type": "cellIs", "operator": "greaterThan", "formula": "5", "fill": "FFFF0000"},
    )
    out = str(tmp_path / "cf.xlsx")
    wb.save(out)

    rules = _rules(out)
    assert len(rules) == 1
    assert rules[0].type == "cellIs"
    assert rules[0].operator == "greaterThan"


def test_color_scale_and_databar(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws.add_conditional_formatting("A1:A10", {"type": "colorScale", "preset": "red_yellow_green"})
    ws.add_conditional_formatting("B1:B10", {"type": "dataBar", "color": "638EC6"})
    out = str(tmp_path / "cf2.xlsx")
    wb.save(out)

    types = {r.type for r in _rules(out)}
    assert "colorScale" in types
    assert "dataBar" in types


def test_text_and_duplicate_rules(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    ws.add_conditional_formatting("A1:A5", {"type": "containsText", "text": "x", "fill": "FFFFFF00"})
    ws.add_conditional_formatting("B1:B5", {"type": "duplicateValues"})
    out = str(tmp_path / "cf3.xlsx")
    wb.save(out)
    types = {r.type for r in _rules(out)}
    assert "containsText" in types
    assert "duplicateValues" in types
