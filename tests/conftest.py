"""Pytest configuration and fixtures for rustypyxl tests."""

import pytest

import rustypyxl


@pytest.fixture
def empty_workbook():
    """Create an empty workbook for testing."""
    return rustypyxl.Workbook()


@pytest.fixture
def workbook_with_sheet():
    """Create a workbook with one sheet."""
    wb = rustypyxl.Workbook()
    wb.create_sheet("Test")
    return wb


@pytest.fixture
def temp_xlsx_path(tmp_path):
    """Temporary path for xlsx output."""
    return str(tmp_path / "test_output.xlsx")


# ---------------------------------------------------------------------------
# Generated fixture files
#
# Feature-specific xlsx files are produced with openpyxl (an independent
# implementation) at session start, so the rustypyxl load path is exercised
# against externally-authored files rather than its own output. This replaces
# the old dependency on gitignored test_*.xlsx files in the repo root, which
# never existed in CI and made these tests silently skip there.
# ---------------------------------------------------------------------------


def _build_simple(path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Item"
    ws["B2"] = 42
    wb.save(path)


def _build_formatting(path):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Styled"
    ws["A1"] = "Bold red"
    ws["A1"].font = Font(bold=True, color="FF0000")
    ws["B1"] = "Filled"
    ws["B1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["C1"] = "Centered"
    ws["C1"].alignment = Alignment(horizontal="center")
    ws["D1"] = 0.5
    ws["D1"].number_format = "0.00%"
    wb.save(path)


def _build_formulas(path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulas"
    ws["A1"] = 2
    ws["A2"] = 3
    ws["A3"] = "=SUM(A1:A2)"
    ws["B1"] = '=CONCATENATE("a","b")'
    wb.save(path)


def _build_comments(path):
    import openpyxl
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commented"
    ws["A1"] = "has comment"
    ws["A1"].comment = Comment("first comment", "author")
    ws["B2"] = "also commented"
    ws["B2"].comment = Comment("second comment", "author")
    wb.save(path)


def _build_hyperlinks(path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Links"
    ws["A1"] = "external"
    ws["A1"].hyperlink = "https://example.com/page"
    ws["A2"] = "mail"
    ws["A2"].hyperlink = "mailto:someone@example.com"
    wb.save(path)


def _build_named_ranges(path):
    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Named"
    ws["A1"] = 1
    wb.defined_names["MyRange"] = DefinedName("MyRange", attr_text="Named!$A$1:$B$2")
    wb.save(path)


def _build_protection(path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Protected"
    ws["A1"] = "locked sheet"
    ws.protection.sheet = True
    ws.protection.password = "secret"
    wb.save(path)


def _build_validation(path):
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validated"
    ws["A1"] = "pick one"
    dv = DataValidation(type="list", formula1='"Yes,No,Maybe"', allowBlank=True)
    dv.add("B1:B10")
    ws.add_data_validation(dv)
    wb.save(path)


_FIXTURE_BUILDERS = {
    "simple.xlsx": _build_simple,
    "formatting.xlsx": _build_formatting,
    "formulas.xlsx": _build_formulas,
    "comments.xlsx": _build_comments,
    "hyperlinks.xlsx": _build_hyperlinks,
    "named_ranges.xlsx": _build_named_ranges,
    "protection.xlsx": _build_protection,
    "validation.xlsx": _build_validation,
}

FIXTURE_NAMES = sorted(_FIXTURE_BUILDERS)


@pytest.fixture(scope="session")
def fixtures_dir(tmp_path_factory):
    """Directory of openpyxl-generated feature files, built once per session."""
    d = tmp_path_factory.mktemp("fixtures")
    for name, build in _FIXTURE_BUILDERS.items():
        build(str(d / name))
    return d


@pytest.fixture
def sample_xlsx_path(fixtures_dir):
    """Path to a simple externally-authored xlsx file."""
    return str(fixtures_dir / "simple.xlsx")
