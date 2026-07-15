"""Rich-text runs within a cell survive load/save, and interop with openpyxl
both directions.

Before this, rustypyxl flattened rich-text runs on read (concatenated the
<r><t> runs into one plain string), silently losing per-run formatting.
"""

import openpyxl
import rustypyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


def _authored(path):
    """A file whose A1 is rich text, written by openpyxl (shared strings)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = CellRichText(
        [
            TextBlock(InlineFont(b=True), "Bold"),
            TextBlock(InlineFont(color="FFFF0000", sz=14), "Red"),
            "plain",
        ]
    )
    wb.save(path)


def test_load_openpyxl_rich_text_keeps_runs(tmp_path):
    src = str(tmp_path / "rich.xlsx")
    _authored(src)

    ws = rustypyxl.load_workbook(src)["S"]
    assert ws["A1"].value == "BoldRedplain", "plain value is the concatenation"

    runs = ws["A1"].rich_text
    assert runs is not None, "rich runs must not be flattened away"
    assert [r["text"] for r in runs] == ["Bold", "Red", "plain"]
    assert runs[0]["bold"] is True
    assert runs[1]["color"] == "#FFFF0000"
    assert runs[1]["size"] == 14.0
    # the trailing unformatted run inherits the cell font: text only, no attrs
    assert runs[2] == {"text": "plain"}


def test_rich_text_round_trips_through_rustypyxl(tmp_path):
    src = str(tmp_path / "rich.xlsx")
    _authored(src)
    out = str(tmp_path / "out.xlsx")

    wb = rustypyxl.load_workbook(src)
    wb.save(out)

    runs = rustypyxl.load_workbook(out)["S"]["A1"].rich_text
    assert [r["text"] for r in runs] == ["Bold", "Red", "plain"]
    assert runs[0]["bold"] is True
    assert runs[1]["color"] == "#FFFF0000"


def test_openpyxl_reads_rustypyxl_output_as_rich(tmp_path):
    src = str(tmp_path / "rich.xlsx")
    _authored(src)
    out = str(tmp_path / "out.xlsx")

    rustypyxl.load_workbook(src).save(out)

    value = openpyxl.load_workbook(out, rich_text=True)["S"]["A1"].value
    assert isinstance(value, CellRichText)
    # first block is the bold run
    first = value[0]
    assert isinstance(first, TextBlock)
    assert first.font.b is True
    assert str(first) == "Bold"


def test_plain_string_is_not_rich(workbook_with_sheet):
    ws = workbook_with_sheet.active
    ws["A1"] = "just text"
    assert ws["A1"].rich_text is None
