"""Tests for sheet protection support."""

import openpyxl

import rustypyxl


class TestProtectionRoundtrip:
    """Test protection survives save/load cycle."""

    def test_load_existing_protection(self, fixtures_dir):
        """Load an externally-authored file with sheet protection."""
        wb = rustypyxl.load_workbook(str(fixtures_dir / "protection.xlsx"))
        assert wb.sheetnames == ["Protected"]
        assert wb["Protected"]["A1"].value == "locked sheet"

    def test_protection_survives_roundtrip(self, fixtures_dir, temp_xlsx_path):
        """Protection (including the password hash) must survive load+save."""
        wb = rustypyxl.load_workbook(str(fixtures_dir / "protection.xlsx"))
        wb.save(temp_xlsx_path)

        chk = openpyxl.load_workbook(temp_xlsx_path)
        prot = chk["Protected"].protection
        assert prot.sheet is True, "sheet protection stripped on round-trip"
        assert prot.password, "password hash stripped on round-trip"
