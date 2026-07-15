"""Charts created through rustypyxl round-trip: a saved workbook opens in
openpyxl with the series, category labels, title, and legend intact.

Charts are a write-path feature -- rustypyxl serializes a chart on save but does
not read charts back on load -- so the reader side of these tests is openpyxl.
"""

import openpyxl
import pytest
import rustypyxl


def _sheet_with_data(wb, name="Data"):
    ws = wb.create_sheet(name)
    wb.write_rows(name, [["Q1", 10], ["Q2", 25], ["Q3", 18], ["Q4", 30]])
    return ws


def test_column_chart_opens_in_openpyxl(tmp_path):
    wb = rustypyxl.Workbook()
    ws = _sheet_with_data(wb)
    ws.add_chart(
        "column",
        "Data!$B$1:$B$4",
        anchor="D1",
        title="Quarterly",
        categories="Data!$A$1:$A$4",
    )
    out = str(tmp_path / "chart.xlsx")
    wb.save(out)

    charts = openpyxl.load_workbook(out)["Data"]._charts
    assert len(charts) == 1
    chart = charts[0]
    # openpyxl models a column chart as a BarChart with type "col"
    assert chart.type == "col"
    assert len(chart.series) == 1
    assert chart.title is not None


def test_multiple_series_from_dicts(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    wb.write_rows("S", [["Jan", 1, 4], ["Feb", 2, 5], ["Mar", 3, 6]])
    ws.add_chart(
        "line",
        [
            {"values": "S!$B$1:$B$3", "name": "A", "categories": "S!$A$1:$A$3"},
            {"values": "S!$C$1:$C$3", "name": "B", "categories": "S!$A$1:$A$3"},
        ],
        anchor="E1",
    )
    out = str(tmp_path / "multi.xlsx")
    wb.save(out)

    chart = openpyxl.load_workbook(out)["S"]._charts[0]
    assert len(chart.series) == 2


def test_pie_chart(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("P")
    wb.write_rows("P", [["Red", 3], ["Green", 5], ["Blue", 2]])
    ws.add_chart(
        "pie", "P!$B$1:$B$3", anchor="D1", categories="P!$A$1:$A$3", legend=None
    )
    out = str(tmp_path / "pie.xlsx")
    wb.save(out)

    from openpyxl.chart import PieChart

    chart = openpyxl.load_workbook(out)["P"]._charts[0]
    assert isinstance(chart, PieChart)


def test_scatter_chart(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("X")
    wb.write_rows("X", [[1, 2], [2, 4], [3, 9]])
    ws.add_chart(
        "scatter",
        {"values": "X!$B$1:$B$3", "categories": "X!$A$1:$A$3"},
        anchor="D1",
    )
    out = str(tmp_path / "scatter.xlsx")
    wb.save(out)

    from openpyxl.chart import ScatterChart

    chart = openpyxl.load_workbook(out)["X"]._charts[0]
    assert isinstance(chart, ScatterChart)


def test_unknown_chart_type_raises(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    with pytest.raises(ValueError):
        ws.add_chart("hexagon", "S!$A$1:$A$3", anchor="B1")


def test_chart_needs_a_series(tmp_path):
    wb = rustypyxl.Workbook()
    ws = wb.create_sheet("S")
    with pytest.raises(ValueError):
        ws.add_chart("column", [], anchor="B1")
