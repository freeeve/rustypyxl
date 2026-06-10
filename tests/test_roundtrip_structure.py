"""Round-trip fidelity tests: structure must survive load+save.

These are differential tests: openpyxl produces a feature-rich file,
rustypyxl loads and re-saves it, and openpyxl verifies nothing was lost.
"""

import pathlib

import openpyxl
import pytest
from openpyxl.worksheet.datavalidation import DataValidation as OxDV
from openpyxl.worksheet.table import Table as OxTable, TableStyleInfo

import rustypyxl


@pytest.fixture
def rich_source(tmp_path) -> pathlib.Path:
    """A feature-rich workbook produced by openpyxl."""
    path = tmp_path / "source.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Main"
    ws["A1"] = "data"
    ws["A3"] = "link"
    ws["A3"].hyperlink = "https://example.com/x?a=1&b=2"
    ws.freeze_panes = "C3"
    ws.auto_filter.ref = "A1:D20"

    dv = OxDV(type="list", formula1='"Yes,No"', allowBlank=True)
    dv.add("E2:E10")
    ws.add_data_validation(dv)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.print_options.gridLines = True

    ws["H1"] = "H1"
    ws["I1"] = "H2"
    ws["H2"] = "a"
    ws["I2"] = 2
    ws["H3"] = "b"
    ws["I3"] = 3
    tab = OxTable(displayName="OxTable1", ref="H1:I3")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)

    hidden = wb.create_sheet("Hidden")
    hidden["A1"] = "secret"
    hidden.sheet_state = "hidden"
    wb.save(path)
    return path


@pytest.fixture
def resaved(rich_source, tmp_path) -> openpyxl.Workbook:
    """The source after a rustypyxl load+save cycle, reopened with openpyxl."""
    out = tmp_path / "resaved.xlsx"
    rwb = rustypyxl.load_workbook(str(rich_source))
    rwb.save(str(out))
    return openpyxl.load_workbook(out)


class TestStructureSurvivesRoundtrip:
    def test_cell_data(self, resaved):
        assert resaved["Main"]["A1"].value == "data"
        assert resaved["Hidden"]["A1"].value == "secret"

    def test_external_hyperlink(self, resaved):
        link = resaved["Main"]["A3"].hyperlink
        assert link is not None, "external hyperlink stripped on round-trip"
        assert link.target == "https://example.com/x?a=1&b=2"

    def test_freeze_panes(self, resaved):
        assert resaved["Main"].freeze_panes == "C3"

    def test_autofilter(self, resaved):
        assert resaved["Main"].auto_filter.ref == "A1:D20"

    def test_data_validation(self, resaved):
        rules = resaved["Main"].data_validations.dataValidation
        assert len(rules) == 1
        rule = rules[0]
        assert rule.type == "list"
        assert rule.formula1 == '"Yes,No"'
        assert str(rule.sqref) == "E2:E10"

    def test_page_setup(self, resaved):
        ws = resaved["Main"]
        assert ws.page_setup.orientation == "landscape"
        assert int(ws.page_setup.paperSize) == 9
        assert ws.print_options.gridLines

    def test_table(self, resaved):
        tables = {t.displayName: t for t in resaved["Main"].tables.values()}
        assert "OxTable1" in tables, "table stripped on round-trip"
        assert tables["OxTable1"].ref == "H1:I3"
        assert tables["OxTable1"].tableStyleInfo.name == "TableStyleMedium9"

    def test_hidden_sheet_state(self, resaved):
        assert resaved["Hidden"].sheet_state == "hidden", (
            "hidden sheet became visible on round-trip"
        )
        assert resaved["Main"].sheet_state == "visible"


class TestStructureApi:
    def test_sheet_state_property(self, rich_source):
        wb = rustypyxl.load_workbook(str(rich_source))
        assert wb["Hidden"].sheet_state == "hidden"
        assert wb["Main"].sheet_state == "visible"

        wb["Hidden"].sheet_state = "visible"
        assert wb["Hidden"].sheet_state == "visible"
        with pytest.raises(ValueError):
            wb["Main"].sheet_state = "invisible"

    def test_active_sheet_respects_loaded_tab(self, tmp_path):
        path = tmp_path / "active.xlsx"
        owb = openpyxl.Workbook()
        owb.create_sheet("Second")
        owb.active = 1
        owb.save(path)

        wb = rustypyxl.load_workbook(str(path))
        assert wb.active.title == "Second"

    def test_rustypyxl_written_hyperlink_works_in_openpyxl(self, tmp_path):
        path = tmp_path / "links.xlsx"
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("S")
        ws["A1"] = "site"
        wb.set_cell_hyperlink("S", 1, 1, "https://example.org/")
        wb.save(str(path))

        chk = openpyxl.load_workbook(path)
        link = chk["S"]["A1"].hyperlink
        assert link is not None
        assert link.target == "https://example.org/"


class TestConditionalFormattingRoundtrip:
    """Conditional formatting + dxfs survive a rustypyxl load+save cycle."""

    @pytest.fixture
    def cf_resaved(self, tmp_path):
        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
        from openpyxl.styles import Font, PatternFill

        src = tmp_path / "cf_src.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CF"
        for r in range(1, 11):
            ws.cell(row=r, column=1, value=r * 10)
        ws.conditional_formatting.add(
            "A1:A10",
            CellIsRule(
                operator="greaterThan",
                formula=["50"],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                font=Font(color="9C0006", bold=True),
            ),
        )
        ws.conditional_formatting.add(
            "A1:A10",
            ColorScaleRule(
                start_type="min", start_color="F8696B", end_type="max", end_color="63BE7B"
            ),
        )
        wb.save(src)

        out = tmp_path / "cf_resaved.xlsx"
        rustypyxl.load_workbook(str(src)).save(str(out))
        return openpyxl.load_workbook(out)

    def test_rules_survive(self, cf_resaved):
        rules = [r for rng in cf_resaved["CF"].conditional_formatting for r in rng.rules]
        types = sorted(r.type for r in rules)
        assert types == ["cellIs", "colorScale"], f"rules lost: {types}"

    def test_dxf_formatting_applies(self, cf_resaved):
        rules = [r for rng in cf_resaved["CF"].conditional_formatting for r in rng.rules]
        cell_is = next(r for r in rules if r.type == "cellIs")
        assert cell_is.operator == "greaterThan"
        assert cell_is.formula == ["50"]
        dxf = cell_is.dxf
        assert dxf is not None, "dxfId not written - rule applies no formatting"
        assert dxf.fill.bgColor.rgb.endswith("FFC7CE")
        assert dxf.font.color.rgb.endswith("9C0006")
        assert dxf.font.b is True

    def test_color_scale_survives(self, cf_resaved):
        rules = [r for rng in cf_resaved["CF"].conditional_formatting for r in rng.rules]
        scale = next(r for r in rules if r.type == "colorScale")
        colors = [c.rgb for c in scale.colorScale.color]
        assert colors[0].endswith("F8696B")
        assert colors[-1].endswith("63BE7B")


class TestNamesFormulasHeadersComments:
    """Defined-name scope, cached formula values, headers/footers, and
    comments survive a rustypyxl load+save cycle."""

    @pytest.fixture
    def full_resaved(self, tmp_path):
        import zipfile

        from openpyxl.comments import Comment
        from openpyxl.workbook.defined_name import DefinedName

        src = tmp_path / "src.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Main"
        ws["A1"] = 2
        ws["B1"] = 3
        ws["C1"] = "=A1+B1"
        ws["A3"].comment = Comment("note text", "author")
        ws.oddHeader.left.text = "Confidential"
        ws.oddFooter.center.text = "Page footer"
        wb.defined_names["GlobalN"] = DefinedName("GlobalN", attr_text="Main!$A$1")
        ws.defined_names["LocalN"] = DefinedName("LocalN", attr_text="Main!$B$1")
        wb.save(src)

        # openpyxl never calculates, so inject the cached formula result the
        # way Excel would have written it
        patched = tmp_path / "patched.xlsx"
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(patched, "w") as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if item == "xl/worksheets/sheet1.xml":
                    data = data.replace(b"<f>A1+B1</f>", b"<f>A1+B1</f><v>5</v>")
                zout.writestr(item, data)

        out = tmp_path / "resaved.xlsx"
        rustypyxl.load_workbook(str(patched)).save(str(out))
        return out

    def test_comment_survives_and_is_displayable(self, full_resaved):
        import zipfile

        chk = openpyxl.load_workbook(full_resaved)
        comment = chk["Main"]["A3"].comment
        assert comment is not None, "comment stripped on round-trip"
        assert comment.text == "note text"
        # The VML part + legacyDrawing are what make Excel display the box
        with zipfile.ZipFile(full_resaved) as z:
            assert "xl/drawings/vmlDrawing1.vml" in z.namelist(), "VML part missing"
            sheet = z.read("xl/worksheets/sheet1.xml").decode()
            assert "<legacyDrawing" in sheet

    def test_header_footer(self, full_resaved):
        m = openpyxl.load_workbook(full_resaved)["Main"]
        assert m.oddHeader.left.text == "Confidential"
        assert m.oddFooter.center.text == "Page footer"

    def test_defined_names_keep_scope(self, full_resaved):
        chk = openpyxl.load_workbook(full_resaved)
        assert "GlobalN" in chk.defined_names, "global defined name lost"
        assert "LocalN" in chk["Main"].defined_names, (
            "sheet-scoped name lost or became global"
        )
        assert "LocalN" not in chk.defined_names

    def test_cached_formula_value(self, full_resaved):
        assert openpyxl.load_workbook(full_resaved)["Main"]["C1"].value == "=A1+B1"
        cached = openpyxl.load_workbook(full_resaved, data_only=True)["Main"]["C1"].value
        assert cached == 5, f"cached formula value lost: {cached!r}"
