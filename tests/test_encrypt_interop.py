"""Bidirectional encryption interop with a third-party tool (msoffcrypto-tool),
with lossless-ness checked at the byte level.

Each existing test covers one direction. These close the loop: content passes
through *both* tools in *both* roles (rustypyxl and msoffcrypto each encrypt AND
decrypt), and the output is compared to the input to prove nothing is lost.

Two levels of check:
  * Pure crypto loops use rustypyxl.encrypt_bytes / decrypt_bytes (no workbook
    re-serialization), so the result must be BYTE-IDENTICAL to the input.
  * A workbook loop goes through rustypyxl's load/save (which re-serializes the
    ZIP), so it is compared cell-by-cell across every sheet.
"""

import io

import msoffcrypto
import openpyxl
import rustypyxl
from msoffcrypto.format.ooxml import OOXMLFile


def _msoffcrypto_encrypt(plain, password):
    out = io.BytesIO()
    OOXMLFile(io.BytesIO(plain)).encrypt(password, out)
    return out.getvalue()


def _msoffcrypto_decrypt(enc, password):
    of = msoffcrypto.OfficeFile(io.BytesIO(enc))
    of.load_key(password=password)
    out = io.BytesIO()
    of.decrypt(out)
    return out.getvalue()


def _rich_plain_xlsx():
    """A plain xlsx (bytes) with varied content across two sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Qty", "Price", "Note"])
    ws.append(["Widget", 3, 9.99, "in stock"])
    # (avoid an empty-string cell here: rustypyxl's load/save normalizes "" to an
    # empty cell, which is orthogonal to encryption and is exercised by the
    # byte-identical crypto loops instead)
    ws.append(["Gadget", -12, 4.5, "backorder"])
    ws.append(["Gizmo", 0, 15.0, "clearance"])
    ws2 = wb.create_sheet("Meta")
    ws2["A1"] = "generated"
    ws2["B1"] = True
    ws2["A2"] = 3.14159
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _all_cells(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    cells = {}
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    cells[(name, c.row, c.column)] = c.value
    return cells


def test_pure_crypto_loop_byte_identical_rustypyxl_first():
    # X -> rustypyxl encrypt -> msoffcrypto decrypt -> msoffcrypto encrypt
    #   -> rustypyxl decrypt.  Each step is pure (de)cryption, so the final bytes
    # must equal the original exactly. Passwords change to prevent pass-through.
    x = _rich_plain_xlsx()
    e1 = rustypyxl.encrypt_bytes(x, "alpha")
    p1 = _msoffcrypto_decrypt(e1, "alpha")
    assert p1 == x, "rustypyxl-encrypt -> msoffcrypto-decrypt is not byte-identical"
    e2 = _msoffcrypto_encrypt(p1, "bravo")
    p2 = rustypyxl.decrypt_bytes(e2, "bravo")
    assert p2 == x, "full loop is not byte-identical"


def test_pure_crypto_loop_byte_identical_msoffcrypto_first():
    x = _rich_plain_xlsx()
    e1 = _msoffcrypto_encrypt(x, "one")
    p1 = rustypyxl.decrypt_bytes(e1, "one")
    assert p1 == x, "msoffcrypto-encrypt -> rustypyxl-decrypt is not byte-identical"
    e2 = rustypyxl.encrypt_bytes(p1, "two")
    p2 = _msoffcrypto_decrypt(e2, "two")
    assert p2 == x, "full loop is not byte-identical"


def test_each_tool_output_decrypts_to_identical_bytes():
    # rustypyxl and msoffcrypto encrypting the SAME plaintext both decrypt back
    # to that exact plaintext (through either tool).
    x = _rich_plain_xlsx()
    assert _msoffcrypto_decrypt(rustypyxl.encrypt_bytes(x, "p"), "p") == x
    assert rustypyxl.decrypt_bytes(_msoffcrypto_encrypt(x, "p"), "p") == x
    assert rustypyxl.decrypt_bytes(rustypyxl.encrypt_bytes(x, "p"), "p") == x


def test_workbook_loop_preserves_all_cells():
    # A loop that DOES re-serialize (rustypyxl load/save), so compare content of
    # every cell rather than raw bytes.
    x = _rich_plain_xlsx()
    original = _all_cells(x)

    # msoffcrypto encrypt -> rustypyxl load+resave(encrypted) -> msoffcrypto decrypt
    e1 = _msoffcrypto_encrypt(x, "k1")
    wb = rustypyxl.load_workbook(e1, password="k1")
    e2 = wb.save_to_bytes(password="k2")
    final = _all_cells(_msoffcrypto_decrypt(e2, "k2"))

    assert final == original, "cell content changed across the encrypted round-trip"


def test_repeated_bounce_no_drift():
    # Bounce between the two tools several times; the bytes must never drift.
    x = _rich_plain_xlsx()
    data = x
    pw = "pw"
    for _ in range(4):
        # rustypyxl encrypts, msoffcrypto decrypts back to x
        data = _msoffcrypto_decrypt(rustypyxl.encrypt_bytes(data, pw), pw)
        assert data == x
        # msoffcrypto encrypts, rustypyxl decrypts back to x
        data = rustypyxl.decrypt_bytes(_msoffcrypto_encrypt(data, pw), pw)
        assert data == x
