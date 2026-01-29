#!/usr/bin/env python3
"""Memory benchmark comparing rustypyxl, calamine, and openpyxl."""

import os
import sys
import tempfile
import tracemalloc
import gc

def get_peak_memory_mb():
    """Get peak memory usage in MB."""
    current, peak = tracemalloc.get_traced_memory()
    return peak / 1024 / 1024

def create_test_file(path, rows, cols, data_type="mixed"):
    """Create a test Excel file using openpyxl."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            if data_type == "numeric":
                ws.cell(row=r, column=c, value=r * c * 1.5)
            elif data_type == "string":
                ws.cell(row=r, column=c, value=f"Cell_{r}_{c}")
            else:  # mixed
                if c % 2 == 0:
                    ws.cell(row=r, column=c, value=r * c * 1.5)
                else:
                    ws.cell(row=r, column=c, value=f"Cell_{r}_{c}")

    wb.save(path)
    wb.close()

def benchmark_read_rustypyxl(path):
    """Benchmark rustypyxl read memory."""
    import rustypyxl

    gc.collect()
    tracemalloc.start()

    wb = rustypyxl.load_workbook(path)
    # Access all data to ensure it's loaded
    data = wb.read_rows('Sheet', min_row=1, max_row=100000)

    peak = get_peak_memory_mb()
    tracemalloc.stop()

    return peak, len(data)

def benchmark_read_calamine(path):
    """Benchmark calamine read memory."""
    try:
        from python_calamine import CalamineWorkbook
    except ImportError:
        return None, 0

    gc.collect()
    tracemalloc.start()

    wb = CalamineWorkbook.from_path(path)
    sheet = wb.get_sheet_by_index(0)
    data = sheet.to_python()

    peak = get_peak_memory_mb()
    tracemalloc.stop()

    return peak, len(data)

def benchmark_read_openpyxl(path):
    """Benchmark openpyxl read memory."""
    import openpyxl

    gc.collect()
    tracemalloc.start()

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    wb.close()

    peak = get_peak_memory_mb()
    tracemalloc.stop()

    return peak, len(data)

def benchmark_write_rustypyxl(path, rows, cols):
    """Benchmark rustypyxl write memory."""
    import rustypyxl

    gc.collect()
    tracemalloc.start()

    wb = rustypyxl.Workbook()
    wb.create_sheet("Sheet")

    data = [[f"R{r}C{c}" if c % 2 else r * c for c in range(cols)] for r in range(rows)]
    wb.write_rows("Sheet", data)
    wb.save(path)

    peak = get_peak_memory_mb()
    tracemalloc.stop()

    return peak

def benchmark_write_openpyxl(path, rows, cols):
    """Benchmark openpyxl write memory."""
    import openpyxl

    gc.collect()
    tracemalloc.start()

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Sheet")

    for r in range(rows):
        row = [f"R{r}C{c}" if c % 2 else r * c for c in range(cols)]
        ws.append(row)
    wb.save(path)

    peak = get_peak_memory_mb()
    tracemalloc.stop()

    return peak

def main():
    print("Memory Benchmark: rustypyxl vs calamine vs openpyxl")
    print("=" * 60)

    # Test configurations
    configs = [
        (10000, 20, "mixed"),
        (50000, 20, "mixed"),
        (100000, 20, "mixed"),
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        # Read benchmarks
        print("\n## Read Memory Usage\n")
        print("| Dataset | rustypyxl | calamine | openpyxl |")
        print("|---------|-----------|----------|----------|")

        for rows, cols, dtype in configs:
            path = os.path.join(tmpdir, f"test_{rows}_{cols}.xlsx")

            # Create test file
            print(f"Creating {rows}x{cols} test file...", file=sys.stderr)
            create_test_file(path, rows, cols, dtype)

            # Benchmark each library
            rusty_mem, _ = benchmark_read_rustypyxl(path)
            cal_mem, _ = benchmark_read_calamine(path)
            opx_mem, _ = benchmark_read_openpyxl(path)

            cal_str = f"{cal_mem:.1f} MB" if cal_mem else "N/A"
            print(f"| {rows//1000}k × {cols} | {rusty_mem:.1f} MB | {cal_str} | {opx_mem:.1f} MB |")

        # Write benchmarks
        print("\n## Write Memory Usage\n")
        print("| Dataset | rustypyxl | openpyxl |")
        print("|---------|-----------|----------|")

        for rows, cols, _ in configs:
            path_rusty = os.path.join(tmpdir, f"write_rusty_{rows}.xlsx")
            path_opx = os.path.join(tmpdir, f"write_opx_{rows}.xlsx")

            rusty_mem = benchmark_write_rustypyxl(path_rusty, rows, cols)
            opx_mem = benchmark_write_openpyxl(path_opx, rows, cols)

            print(f"| {rows//1000}k × {cols} | {rusty_mem:.1f} MB | {opx_mem:.1f} MB |")

if __name__ == "__main__":
    main()
