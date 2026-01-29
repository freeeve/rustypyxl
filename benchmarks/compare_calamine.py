"""Benchmarks comparing rustypyxl vs calamine vs openpyxl read performance.

Calamine is a high-performance Rust Excel reader.
"""

import time
import tempfile
import os

import openpyxl
import python_calamine
import rustypyxl


def create_test_file(rows: int, cols: int, use_strings: bool = False) -> str:
    """Create a test xlsx file with openpyxl."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name

    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            if use_strings:
                ws.cell(r, c).value = f"R{r}C{c}"
            else:
                ws.cell(r, c).value = r * c
    wb.save(path)
    return path


def benchmark_read(path: str, rows: int, cols: int):
    """Benchmark reading a file with all three libraries."""
    cells = rows * cols

    # openpyxl
    start = time.perf_counter()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # Force reading all cells
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    wb.close()
    openpyxl_time = time.perf_counter() - start

    # calamine
    start = time.perf_counter()
    workbook = python_calamine.CalamineWorkbook.from_path(path)
    sheet_name = workbook.sheet_names[0]
    data = workbook.get_sheet_by_name(sheet_name).to_python()
    calamine_time = time.perf_counter() - start

    # rustypyxl
    start = time.perf_counter()
    wb = rustypyxl.load_workbook(path)
    sheet = wb.active.title
    data = wb.read_rows(sheet)
    rustypyxl_time = time.perf_counter() - start

    return openpyxl_time, calamine_time, rustypyxl_time


def main():
    print("=" * 80)
    print("Read Performance: rustypyxl vs calamine vs openpyxl")
    print("=" * 80)

    test_cases = [
        (1000, 10, False, "1k rows x 10 cols (numeric)"),
        (1000, 10, True, "1k rows x 10 cols (strings)"),
        (10000, 20, False, "10k rows x 20 cols (numeric)"),
        (10000, 20, True, "10k rows x 20 cols (strings)"),
        (50000, 20, False, "50k rows x 20 cols (numeric)"),
        (100000, 20, False, "100k rows x 20 cols (numeric)"),
    ]

    results = []

    for rows, cols, use_strings, desc in test_cases:
        cells = rows * cols
        print(f"\n{desc} ({cells:,} cells)")
        print("-" * 50)

        # Create test file
        print("  Creating test file...", end=" ", flush=True)
        start = time.perf_counter()
        path = create_test_file(rows, cols, use_strings)
        create_time = time.perf_counter() - start
        file_size = os.path.getsize(path) / (1024 * 1024)
        print(f"{create_time:.1f}s, {file_size:.1f} MB")

        # Benchmark
        openpyxl_time, calamine_time, rustypyxl_time = benchmark_read(path, rows, cols)

        print(f"  openpyxl:   {openpyxl_time:.3f}s ({cells/openpyxl_time:,.0f} cells/sec)")
        print(f"  calamine:   {calamine_time:.3f}s ({cells/calamine_time:,.0f} cells/sec)")
        print(f"  rustypyxl:  {rustypyxl_time:.3f}s ({cells/rustypyxl_time:,.0f} cells/sec)")

        # Speedups
        print(f"  rustypyxl vs openpyxl: {openpyxl_time/rustypyxl_time:.1f}x")
        print(f"  rustypyxl vs calamine: {calamine_time/rustypyxl_time:.1f}x")
        print(f"  calamine vs openpyxl:  {openpyxl_time/calamine_time:.1f}x")

        results.append({
            "desc": desc,
            "cells": cells,
            "openpyxl": openpyxl_time,
            "calamine": calamine_time,
            "rustypyxl": rustypyxl_time,
        })

        os.unlink(path)

    # Summary table
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"{'Test Case':<35} {'openpyxl':>10} {'calamine':>10} {'rustypyxl':>10} {'vs calamine':>12}")
    print("-" * 80)

    for r in results:
        vs_calamine = r['calamine'] / r['rustypyxl']
        faster_slower = "faster" if vs_calamine > 1 else "slower"
        print(f"{r['desc']:<35} {r['openpyxl']:>9.3f}s {r['calamine']:>9.3f}s {r['rustypyxl']:>9.3f}s {vs_calamine:>10.1f}x {faster_slower}")

    print("=" * 80)


if __name__ == "__main__":
    main()
