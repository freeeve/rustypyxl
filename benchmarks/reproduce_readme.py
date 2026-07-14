"""Re-measure the README's read/write benchmark claims, honestly.

- Reports MINIMUM wall time across runs, not the mean: the fastest run is the
  one least disturbed by other work, which is the right estimator on a machine
  that is never perfectly idle. (The old scripts averaged.)
- Meant to run under `taskman lock run -max-load N`, which certifies the box was
  quiet enough and exits non-zero otherwise -- so a contaminated run cannot
  silently publish.
- Prints copy-pasteable README table rows.

Requires openpyxl and python-calamine for the comparison.

Usage: python benchmarks/reproduce_readme.py [write|read|all]
"""

import gc
import os
import sys
import tempfile
import time

import openpyxl
import python_calamine
import rustypyxl

TMP = tempfile.mkdtemp(prefix="rustypyxl_bench_")


def best(fn, runs):
    """Minimum wall time over `runs` invocations (with a gc between)."""
    times = []
    for _ in range(runs):
        gc.collect()
        t = time.perf_counter()
        fn()
        times.append(time.perf_counter() - t)
    return min(times)


# ---------------------------------------------------------------- write (1M)

def bench_write():
    rows, cols = 1_000_000, 20
    # Mixed data, matching a realistic sheet: half strings, half numbers.
    data = [[f"R{r}C{c}" if c % 2 else r * c for c in range(cols)] for r in range(rows)]
    print(f"# write {rows:,} x {cols} ({rows*cols:,} cells), mixed", flush=True)

    def rusty():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        wb.write_rows(ws.title, data, start_row=1, start_col=1)
        wb.save(os.path.join(TMP, "w_rusty.xlsx"))

    def rusty_stream():
        wb = rustypyxl.WriteOnlyWorkbook(os.path.join(TMP, "w_stream.xlsx"))
        wb.create_sheet("Data")
        # append_rows: batch path, GIL released per batch
        for i in range(0, len(data), 10_000):
            wb.append_rows(data[i:i + 10_000])
        wb.close()

    def op():
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Data")
        for row in data:
            ws.append(row)
        wb.save(os.path.join(TMP, "w_op.xlsx"))

    # rustypyxl is fast: 3 runs. openpyxl write_only ~minutes: 1 run.
    r = best(rusty, 3)
    s = best(rusty_stream, 3)
    o = best(op, 1)
    print(f"| rustypyxl (write_rows) | {r:.1f}s |")
    print(f"| rustypyxl (WriteOnlyWorkbook) | {s:.1f}s |")
    print(f"| openpyxl (write_only) | {o:.1f}s |")
    print(f"# speedup vs openpyxl: write_rows {o/r:.0f}x, streaming {o/s:.0f}x", flush=True)


# ------------------------------------------------------------------- read

def make_fixture(rows, cols, kind):
    path = os.path.join(TMP, f"r_{rows}_{kind}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            if kind == "numeric":
                ws.cell(r, c).value = r * c
            elif kind == "strings":
                ws.cell(r, c).value = f"R{r}C{c}"
            else:  # mixed
                ws.cell(r, c).value = f"R{r}C{c}" if c % 2 else r * c
    wb.save(path)
    return path


def bench_read():
    cases = [
        (10_000, 20, "numeric"),
        (10_000, 20, "strings"),
        (100_000, 20, "numeric"),
        (100_000, 20, "mixed"),
    ]
    print("\n# read (min wall time)")
    print("| Dataset | rustypyxl | calamine | openpyxl |")
    print("|---------|-----------|----------|----------|")
    for rows, cols, kind in cases:
        path = make_fixture(rows, cols, kind)

        def rusty():
            wb = rustypyxl.load_workbook(path)
            _ = wb.read_rows(wb.active.title)

        def cal():
            wb = python_calamine.CalamineWorkbook.from_path(path)
            _ = wb.get_sheet_by_index(0).to_python()

        def op():
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            _ = [[c.value for c in row] for row in wb.active.iter_rows()]
            wb.close()

        r = best(rusty, 3)
        c = best(cal, 3)
        o = best(op, 2)  # openpyxl read is slow; two runs is enough for a min
        os.unlink(path)
        print(f"| {rows//1000}k x {cols} ({kind}) | {r:.2f}s | {c:.2f}s | {o:.2f}s |", flush=True)


if __name__ == "__main__":
    what = sys.argv[1] if len(sys.argv) > 1 else "all"
    print(f"openpyxl {openpyxl.__version__}, python-calamine present", flush=True)
    if what in ("write", "all"):
        bench_write()
    if what in ("read", "all"):
        bench_read()
