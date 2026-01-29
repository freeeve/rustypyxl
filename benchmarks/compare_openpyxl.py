"""Benchmarks comparing rustypyxl vs openpyxl performance.

Usage:
    python benchmarks/compare_openpyxl.py

Requires openpyxl to be installed for comparison.
"""

import time
import tempfile
import os
from contextlib import contextmanager
from typing import Callable, Any

# Try to import both libraries
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed, comparison benchmarks will be skipped")

import rustypyxl


@contextmanager
def timer(label: str):
    """Context manager for timing code blocks."""
    start = time.perf_counter()
    yield
    elapsed = time.perf_counter() - start
    print(f"  {label}: {elapsed:.3f}s")


def benchmark(func: Callable, iterations: int = 3) -> float:
    """Run a function multiple times and return average time."""
    times = []
    for _ in range(iterations):
        start = time.perf_counter()
        func()
        times.append(time.perf_counter() - start)
    return sum(times) / len(times)


class BenchmarkResults:
    """Collect and display benchmark results."""

    def __init__(self):
        self.results = []

    def add(self, name: str, rustypyxl_time: float, openpyxl_time: float | None):
        self.results.append((name, rustypyxl_time, openpyxl_time))

    def print_summary(self):
        print("\n" + "=" * 70)
        print("BENCHMARK SUMMARY")
        print("=" * 70)
        print(f"{'Benchmark':<35} {'rustypyxl':>10} {'openpyxl':>10} {'Speedup':>10}")
        print("-" * 70)

        for name, rust_time, openpyxl_time in self.results:
            rust_str = f"{rust_time:.3f}s"
            if openpyxl_time is not None:
                openpyxl_str = f"{openpyxl_time:.3f}s"
                speedup = openpyxl_time / rust_time if rust_time > 0 else float('inf')
                speedup_str = f"{speedup:.1f}x"
            else:
                openpyxl_str = "N/A"
                speedup_str = "N/A"

            print(f"{name:<35} {rust_str:>10} {openpyxl_str:>10} {speedup_str:>10}")

        print("=" * 70)


results = BenchmarkResults()


def bench_create_workbook_small():
    """Benchmark: Create workbook with 100 rows x 10 columns."""
    rows, cols = 100, 10

    def rustypyxl_impl():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        sheet = ws.title
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                wb.set_cell_value(sheet, row, col, f"R{row}C{col}")
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    def openpyxl_impl():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = f"R{row}C{col}"
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    print(f"\nCreate & Save: {rows} rows x {cols} cols ({rows * cols:,} cells)")
    rust_time = benchmark(rustypyxl_impl)
    openpyxl_time = benchmark(openpyxl_impl) if HAS_OPENPYXL else None
    results.add(f"Create small ({rows}x{cols})", rust_time, openpyxl_time)


def bench_create_workbook_medium():
    """Benchmark: Create workbook with 1000 rows x 20 columns."""
    rows, cols = 1000, 20

    def rustypyxl_impl():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        sheet = ws.title
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                wb.set_cell_value(sheet, row, col, f"R{row}C{col}")
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    def openpyxl_impl():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = f"R{row}C{col}"
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    print(f"\nCreate & Save: {rows} rows x {cols} cols ({rows * cols:,} cells)")
    rust_time = benchmark(rustypyxl_impl)
    openpyxl_time = benchmark(openpyxl_impl) if HAS_OPENPYXL else None
    results.add(f"Create medium ({rows}x{cols})", rust_time, openpyxl_time)


def bench_create_workbook_large():
    """Benchmark: Create workbook with 10000 rows x 20 columns."""
    rows, cols = 10000, 20

    def rustypyxl_impl():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        sheet = ws.title
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                wb.set_cell_value(sheet, row, col, f"R{row}C{col}")
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    def openpyxl_impl():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = f"R{row}C{col}"
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    print(f"\nCreate & Save: {rows} rows x {cols} cols ({rows * cols:,} cells)")
    rust_time = benchmark(rustypyxl_impl, iterations=1)
    openpyxl_time = benchmark(openpyxl_impl, iterations=1) if HAS_OPENPYXL else None
    results.add(f"Create large ({rows}x{cols})", rust_time, openpyxl_time)


def bench_bulk_write_openpyxl_comparison():
    """Benchmark: Compare bulk write APIs - rustypyxl vs openpyxl write_only mode."""
    rows, cols = 10000, 20

    # Pre-generate data
    data = [[f"R{r}C{c}" for c in range(cols)] for r in range(rows)]

    def rustypyxl_bulk():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        wb.write_rows(ws.title, data, start_row=1, start_col=1)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    def openpyxl_write_only():
        # openpyxl's optimized write_only mode with append()
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Data")
        for row in data:
            ws.append(row)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    print(f"\nBulk Write Comparison: {rows} rows x {cols} cols ({rows * cols:,} cells)")
    print("  (rustypyxl write_rows vs openpyxl write_only + append)")

    rust_time = benchmark(rustypyxl_bulk, iterations=3)
    openpyxl_time = benchmark(openpyxl_write_only, iterations=3) if HAS_OPENPYXL else None

    if openpyxl_time:
        speedup = openpyxl_time / rust_time
        print(f"  rustypyxl (write_rows):      {rust_time:.3f}s")
        print(f"  openpyxl (write_only+append): {openpyxl_time:.3f}s")
        print(f"  Speedup: {speedup:.1f}x")

    results.add(f"Bulk write vs write_only", rust_time, openpyxl_time)


def bench_load_and_read():
    """Benchmark: Load existing file and read all cells."""
    # First create a test file
    rows, cols = 1000, 10

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name

    # Create with openpyxl (for cross-compatibility)
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = row * col
        wb.save(temp_path)
    else:
        # Fallback to rustypyxl
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = row * col
        wb.save(temp_path)

    def rustypyxl_impl():
        wb = rustypyxl.load_workbook(temp_path)
        sheet = wb.active.title
        total = 0
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                val = wb.get_cell_value(sheet, row, col)
                if val is not None:
                    total += 1
        return total

    def openpyxl_impl():
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active
        total = 0
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                val = ws.cell(row, col).value
                if val is not None:
                    total += 1
        return total

    print(f"\nLoad & Read: {rows} rows x {cols} cols ({rows * cols:,} cells)")
    rust_time = benchmark(rustypyxl_impl)
    openpyxl_time = benchmark(openpyxl_impl) if HAS_OPENPYXL else None
    results.add(f"Load & read ({rows}x{cols})", rust_time, openpyxl_time)

    os.unlink(temp_path)


def bench_modify_existing():
    """Benchmark: Load existing file, modify cells, and save."""
    rows, cols = 500, 10

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        output_path = f.name

    # Create initial file with openpyxl for cross-compatibility
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = row * col
        wb.save(temp_path)
    else:
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = row * col
        wb.save(temp_path)

    def rustypyxl_impl():
        wb = rustypyxl.load_workbook(temp_path)
        sheet = wb.active.title
        # Modify every other row
        for row in range(1, rows + 1, 2):
            for col in range(1, cols + 1):
                wb.set_cell_value(sheet, row, col, "modified")
        wb.save(output_path)

    def openpyxl_impl():
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active
        # Modify every other row
        for row in range(1, rows + 1, 2):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = "modified"
        wb.save(output_path)

    print(f"\nLoad, Modify, Save: {rows} rows x {cols} cols")
    rust_time = benchmark(rustypyxl_impl)
    openpyxl_time = benchmark(openpyxl_impl) if HAS_OPENPYXL else None
    results.add(f"Modify existing ({rows}x{cols})", rust_time, openpyxl_time)

    os.unlink(temp_path)
    os.unlink(output_path)


def bench_append_rows():
    """Benchmark: Append rows to existing workbook (common data loading pattern)."""
    initial_rows = 100
    append_rows = 500
    cols = 10

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        output_path = f.name

    # Create initial file with openpyxl for cross-compatibility
    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, initial_rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = f"initial_{row}_{col}"
        wb.save(temp_path)
    else:
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        for row in range(1, initial_rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = f"initial_{row}_{col}"
        wb.save(temp_path)

    def rustypyxl_impl():
        wb = rustypyxl.load_workbook(temp_path)
        sheet = wb.active.title
        # Append new rows
        start_row = initial_rows + 1
        for row in range(start_row, start_row + append_rows):
            for col in range(1, cols + 1):
                wb.set_cell_value(sheet, row, col, f"new_{row}_{col}")
        wb.save(output_path)

    def openpyxl_impl():
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active
        # Append new rows
        start_row = initial_rows + 1
        for row in range(start_row, start_row + append_rows):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = f"new_{row}_{col}"
        wb.save(output_path)

    print(f"\nAppend {append_rows} rows to existing ({initial_rows} rows)")
    rust_time = benchmark(rustypyxl_impl)
    openpyxl_time = benchmark(openpyxl_impl) if HAS_OPENPYXL else None
    results.add(f"Append rows ({append_rows} rows)", rust_time, openpyxl_time)

    os.unlink(temp_path)
    os.unlink(output_path)


def bench_numeric_data():
    """Benchmark: Write numeric data (common for data analysis)."""
    rows, cols = 5000, 10

    def rustypyxl_impl():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Numbers")
        sheet = ws.title
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                wb.set_cell_value(sheet, row, col, row * col * 0.123)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    def openpyxl_impl():
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = row * col * 0.123
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    print(f"\nNumeric data: {rows} rows x {cols} cols ({rows * cols:,} cells)")
    rust_time = benchmark(rustypyxl_impl, iterations=1)
    openpyxl_time = benchmark(openpyxl_impl, iterations=1) if HAS_OPENPYXL else None
    results.add(f"Numeric data ({rows}x{cols})", rust_time, openpyxl_time)


def bench_formulas():
    """Benchmark: Write formulas."""
    rows = 1000

    def rustypyxl_impl():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Formulas")
        sheet = ws.title
        # Column A: values, Column B: formulas referencing A
        for row in range(1, rows + 1):
            wb.set_cell_value(sheet, row, 1, row)
            wb.set_cell_value(sheet, row, 2, f"=A{row}*2")
            wb.set_cell_value(sheet, row, 3, f"=SUM(A1:A{row})")
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    def openpyxl_impl():
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in range(1, rows + 1):
            ws.cell(row, 1).value = row
            ws.cell(row, 2).value = f"=A{row}*2"
            ws.cell(row, 3).value = f"=SUM(A1:A{row})"
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    print(f"\nFormulas: {rows} rows with formulas")
    rust_time = benchmark(rustypyxl_impl)
    openpyxl_time = benchmark(openpyxl_impl) if HAS_OPENPYXL else None
    results.add(f"Formulas ({rows} rows)", rust_time, openpyxl_time)


def bench_multiple_sheets():
    """Benchmark: Create workbook with multiple sheets."""
    sheets = 10
    rows, cols = 100, 10

    def rustypyxl_impl():
        wb = rustypyxl.Workbook()
        for i in range(sheets):
            ws = wb.create_sheet(f"Sheet{i}")
            sheet = ws.title
            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    wb.set_cell_value(sheet, row, col, f"S{i}R{row}C{col}")
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    def openpyxl_impl():
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        for i in range(sheets):
            ws = wb.create_sheet(f"Sheet{i}")
            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    ws.cell(row, col).value = f"S{i}R{row}C{col}"
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    print(f"\nMultiple sheets: {sheets} sheets x {rows}x{cols} each")
    rust_time = benchmark(rustypyxl_impl)
    openpyxl_time = benchmark(openpyxl_impl) if HAS_OPENPYXL else None
    results.add(f"Multi-sheet ({sheets} sheets)", rust_time, openpyxl_time)


# ============================================================================
# BULK API BENCHMARKS (rustypyxl-specific optimizations)
# ============================================================================

def bench_bulk_write_rows():
    """Benchmark: Bulk write rows using workbook API."""
    rows_data = [[f"R{r}C{c}" for c in range(1, 21)] for r in range(1, 1001)]

    def rustypyxl_cell_by_cell():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        sheet_name = ws.title
        for row_idx, row_data in enumerate(rows_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                wb.set_cell_value(sheet_name, row_idx, col_idx, value)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    def rustypyxl_bulk():
        wb = rustypyxl.Workbook()
        ws = wb.create_sheet("Data")
        sheet_name = ws.title
        # Use bulk API on workbook
        wb.write_rows(sheet_name, rows_data, start_row=1, start_col=1)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            wb.save(f.name)
            os.unlink(f.name)

    print(f"\nBulk write: 1000 rows x 20 cols (20,000 cells)")

    # Check if bulk API is available
    wb_test = rustypyxl.Workbook()
    has_bulk = hasattr(wb_test, 'write_rows')

    if has_bulk:
        print("  (bulk API available)")
        cell_time = benchmark(rustypyxl_cell_by_cell)
        bulk_time = benchmark(rustypyxl_bulk)
        print(f"  Cell-by-cell: {cell_time:.3f}s")
        print(f"  Bulk API:     {bulk_time:.3f}s")
        print(f"  Bulk speedup: {cell_time/bulk_time:.1f}x")
        results.add("Bulk write (1000x20)", bulk_time, None)
    else:
        print("  (bulk API not yet implemented)")
        cell_time = benchmark(rustypyxl_cell_by_cell)
        print(f"  Cell-by-cell: {cell_time:.3f}s")


def bench_bulk_read_rows():
    """Benchmark: Bulk read rows using workbook API."""
    rows, cols = 1000, 20

    # Create test file with openpyxl
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row, col).value = f"R{row}C{col}"
        wb.save(temp_path)
    else:
        # Skip if no openpyxl
        print("\nBulk read: (skipped - no openpyxl to create test file)")
        return

    def rustypyxl_cell_by_cell():
        wb = rustypyxl.load_workbook(temp_path)
        sheet_name = wb.active.title
        total = 0
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                val = wb.get_cell_value(sheet_name, row, col)
                if val is not None:
                    total += 1
        return total

    def rustypyxl_bulk():
        wb = rustypyxl.load_workbook(temp_path)
        sheet_name = wb.active.title
        data = wb.read_rows(sheet_name)
        return sum(1 for row in data for val in row if val is not None)

    print(f"\nBulk read: {rows} rows x {cols} cols ({rows * cols:,} cells)")

    # Check if bulk API is available
    wb_test = rustypyxl.Workbook()
    has_bulk = hasattr(wb_test, 'read_rows')

    if has_bulk:
        print("  (bulk API available)")
        cell_time = benchmark(rustypyxl_cell_by_cell)
        bulk_time = benchmark(rustypyxl_bulk)
        print(f"  Cell-by-cell: {cell_time:.3f}s")
        print(f"  Bulk API:     {bulk_time:.3f}s")
        print(f"  Bulk speedup: {cell_time/bulk_time:.1f}x")
        results.add("Bulk read (1000x20)", bulk_time, None)
    else:
        print("  (bulk API not yet implemented)")
        cell_time = benchmark(rustypyxl_cell_by_cell)
        print(f"  Cell-by-cell: {cell_time:.3f}s")

    os.unlink(temp_path)


def main():
    print("=" * 70)
    print("rustypyxl vs openpyxl Benchmarks")
    print("=" * 70)

    if HAS_OPENPYXL:
        print(f"openpyxl version: {openpyxl.__version__}")
    print(f"rustypyxl available: Yes")

    # Run benchmarks
    bench_create_workbook_small()
    bench_create_workbook_medium()
    bench_create_workbook_large()
    bench_bulk_write_openpyxl_comparison()  # Fair bulk comparison
    bench_load_and_read()
    bench_modify_existing()
    bench_append_rows()
    bench_numeric_data()
    bench_formulas()
    bench_multiple_sheets()

    # Bulk API benchmarks
    bench_bulk_write_rows()
    bench_bulk_read_rows()

    # Print summary
    results.print_summary()


if __name__ == "__main__":
    main()
